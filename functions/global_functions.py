# Imports
import powergama
import pathlib
import numpy as np
import pandas as pd
import time
import folium
from folium.features import DivIcon
import math
from math import radians, degrees, atan2, cos, sin
import logging

import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import matplotlib.dates as mdates

import powergama.scenarios as pgs
from IPython.display import display
import branca.colormap as cm
from openpyxl.utils.datetime import days_to_time
from powergama.GIS import _pointBetween
from powergama.database import Database  # Import Database-Class specifically
from powergama.GridData import GridData  # Import GridData-Class specifically
from functions.database_functions import *
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from collections import defaultdict
import time



def read_grid_data(version,
                   date_start,
                   date_end,
                   data_path,
                   case,
                   ):
    """
    Reads and processes grid data for a specified year and date range.

    Parameters:
        year (int): The year for which data should be loaded.
        version (str): The version of the dataset to be used.
        date_start (str): The start date of the simulation period in 'YYYY-MM-DD' format.
        date_end (str): The end date of the simulation period in 'YYYY-MM-DD' format.
        data_path (str or pathlib.Path): The base path to the data directory.

    Returns:
        powergama.GridData: An instance of GridData containing the processed grid data.
    """
    # Calculate and print the number of simulation hours and years
    datapath_GridData = data_path / "system"
    file_storval_filling = data_path / f"storage/profiles_storval_filling_{case}_{version}.csv"
    file_30y_profiles = data_path / "timeseries_profiles.csv"

    # Initialize GridData object
    data = powergama.GridData()
    data.readGridData(nodes=datapath_GridData / f"node_{case}_{version}.csv",
                      ac_branches=datapath_GridData / f"branch_{case}_{version}.csv",
                      dc_branches=datapath_GridData / f"dcbranch_{case}_{version}.csv",
                      generators=datapath_GridData / f"generator_{case}_{version}.csv",
                      consumers=datapath_GridData / f"consumer_{case}_{version}.csv")

    # Read and process 30-year profiles
    profiles_30y = pd.read_csv(file_30y_profiles, index_col=0, parse_dates=True)
    profiles_30y["const"] = 1
    data.profiles = profiles_30y[(profiles_30y.index >= date_start) & (profiles_30y.index <= date_end)].reset_index()
    data.storagevalue_time = data.profiles[["const"]]

    # Read storage value filling data
    storval_filling = pd.read_csv(file_storval_filling)
    data.storagevalue_filling = storval_filling

    # Set the timerange and time delta for the simulation
    data.timerange = list(range(data.profiles.shape[0]))
    data.timeDelta = 1.0  # hourly data

    # Calculate and print the number of simulation hours and years
    num_hours = len(data.profiles) # data.timerange[-1] - data.timerange[0]
    print(f'Simulation hours: {num_hours}')
    num_years = num_hours / (365.2425 * 24)
    print(f'Simulation years: {np.round(num_years, 3)}')

    # Filter offshore wind farms by year:
    # data.generator = data.generator[~(data.generator["year"] > year)].reset_index(drop=True)

    # remove zero capacity generators:
    m_gen_hascap = data.generator["pmax"] > 0
    data.generator = data.generator[m_gen_hascap].reset_index(drop=True)

    return data


# Read and configure grid
def setup_grid(version,
               date_start,
               date_end,
               data_path,
               case,
               ):
    """
    Set up grid data and initialize a simulation scenario.

    This function reads grid data for the specified year and date range,
    then configures the base grid data with a specific scenario file.

    Parameters:
        year (int): The year for which data should be loaded.
        version (str): The version of the dataset to be used.
        date_start (str): The start date of the simulation period.
        date_end (str): The end date of the simulation period.

    Returns:
        data (Scenario): Configured grid data for simulation.
        time_max_min (list): List containing the start and end indices for the simulation timeframe.
    """
    print(f"Using version: {version}")
    data = read_grid_data(version, date_start, date_end, data_path, case)
    time_max_min = [0, len(data.timerange)]
    return data, time_max_min



def solve_lp(data,
             sql_file,
             loss_method,
             replace,
             solver):
    """
    Solves a linear programming problem using the given grid data and stores the results in a SQL file.

    Parameters:
        data (powergama.GridData): The grid data to be used for the linear programming problem.
        sql_file (str): The path to the SQL file where the results will be stored.
        loss_method (str): The loss method to be used.
        replace (bool): If True, replace the existing solution with a new solution.
        nuclear_availability (float): The nuclear availability factor to be used.
        week_MSO (int): The week of MSO (Maintenance Start Order) to be used.

    Returns:
        powergama.Results: The results of the linear programming problem.
    """

    lp = powergama.LpProblem(grid=data, lossmethod=loss_method)  # lossmethod; 0=no losses, 1=linearised losses, 2=added as load
    # if replace = False, bruker kun sql_file som input
    res = powergama.Results(data, sql_file, replace=replace)
    if replace:
        start_time = time.time()
        lp.solve(res, solver=solver)
        end_time = time.time()
        print("\nSimulation time = {:.2f} seconds".format(end_time - start_time))
        print("\nSimulation time = {:.2f} minutes".format((end_time - start_time)/60))

    return res




########################################### DATA COLLECTION ######################################################
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
def getPTDF_Matrix(data: GridData):
    # === GET FLOW FACTOR MATRIX ===

    # Compute matrices
    Bbus, DAmatrix = data.compute_power_flow_matrices()

    # Compute PTDF explicitly (correct formula)
    Bbus_inv = np.linalg.pinv(Bbus.todense())
    PTDF = DAmatrix @ Bbus_inv

    # Convert to array
    PTDF_array = PTDF if isinstance(PTDF, np.ndarray) else PTDF.toarray()
    PTDF_df = pd.DataFrame(PTDF_array, columns=data.node.index, index=data.branch.index)

    # Display basic info
    print("Matrix shape:", PTDF_df.shape)
    print(PTDF_df.describe(percentiles=[.25, .5, .75, .95]))

    # Visualize with heatmap
    plt.figure(figsize=(20, 20))
    sns.heatmap(PTDF_array, cmap='RdBu', center=0,
                cbar_kws={'label': 'PTDF Sensitivity'})

    plt.xlabel('Nodes')
    plt.ylabel('Lines')
    plt.title('PTDF Matrix Visualization')
    plt.tight_layout()
    plt.show()



def createZonePriceMatrix(data, database, zones, year_range, TIMEZONE, SIM_YEAR_START, SIM_YEAR_END):
    log_messages = []
    zonal_price_map = pd.DataFrame(index=zones)

    # Set up time range for the entire period
    try:
        START = {"year": year_range[0], "month": 1, "day": 1, "hour": 0}
        END = {"year": year_range[-1], "month": 12, "day": 31, "hour": 23}
        start_datetime = datetime(START["year"], START["month"], START["day"], START["hour"], 0, tzinfo=TIMEZONE)
        time_range = get_hour_range(SIM_YEAR_START, SIM_YEAR_END, TIMEZONE, START, END)
        date_index = pd.date_range(start=start_datetime, periods=time_range[-1], freq='h', inclusive='left')
    except Exception as e:
        log_messages.append(f"❌ Failed to generate time range for {year_range[0]}-{year_range[-1]}: {e}")
        print(f"❌ Failed to generate time range for {year_range[0]}-{year_range[-1]}: {e}")
        return zonal_price_map, "\n".join(log_messages)

    # Fetch all nodal prices for the entire period
    try:
        log_messages.append(f"📡 Fetching all nodal prices for {year_range[0]}-{year_range[-1]}...")
        print(f"📡 Fetching all nodal prices for {year_range[0]}-{year_range[-1]}...")
        start_time = time.time()
        price_data = database.getResultNodalPricesAll(time_range)
        end_time = time.time()
        log_messages.append(f"⏱️ Fetching duration: {end_time - start_time:.2f} seconds.")
        print(f"⏱️ Fetching duration: {end_time - start_time:.2f} seconds.")
    except Exception as e:
        log_messages.append(f"❌ Failed to fetch nodal prices for {year_range[0]}-{year_range[-1]}: {e}")
        print(f"❌ Failed to fetch nodal prices for {year_range[0]}-{year_range[-1]}: {e}")
        return zonal_price_map, "\n".join(log_messages)

    # Convert price data to DataFrame
    try:
        # Create DataFrame from list of (timestep, node_index, nodalprice)
        print("📊 Processing price data...")
        start_time = time.time()
        df_all = pd.DataFrame(price_data, columns=['timestep', 'node_index', 'nodalprice'])
        # Pivot to get nodes as columns, timesteps as rows
        df_pivot = df_all.pivot(index='timestep', columns='node_index', values='nodalprice')
        # Align with date_index
        df_pivot.index = date_index[:len(df_pivot)]
        end_time = time.time()
        log_messages.append(f"⏱️ Processing duration: {end_time - start_time:.2f} seconds.")
        print(f"⏱️ Processing duration: {end_time - start_time:.2f} seconds.")
    except Exception as e:
        log_messages.append(f"❌ Failed to process price data for {year_range[0]}-{year_range[-1]}: {e}")
        print(f"❌ Failed to process price data for {year_range[0]}-{year_range[-1]}: {e}")
        return zonal_price_map, "\n".join(log_messages)

    # Process each year
    for year in year_range:
        try:
            # Define year boundaries
            year_start = datetime(year, 1, 1, 0, 0, tzinfo=TIMEZONE)
            year_end = datetime(year, 12, 31, 23, 0, tzinfo=TIMEZONE) if year == year_range[-1] else datetime(year + 1, 1, 1, 0, 0, tzinfo=TIMEZONE)
            # Filter DataFrame for this year
            df_year = df_pivot.loc[year_start:year_end]

            # Process each zone
            for zone in zones:
                try:
                    nodes_in_zone = data.node[data.node['zone'] == zone].index.tolist()
                    # Exclude nodes ending with 'SINK'
                    nodes_in_zone = [node for node in nodes_in_zone if not data.node.loc[node, 'id'].endswith('SINK')]
                    if not nodes_in_zone:
                        log_messages.append(f"⚠️ No nodes found for zone {zone} — skipping.")
                        print(f"⚠️ No nodes found for zone {zone} — skipping.")
                        continue

                    # Filter DataFrame to only nodes in this zone
                    zone_nodes = [node for node in nodes_in_zone if node in df_year.columns]
                    if not zone_nodes:
                        log_messages.append(f"⚠️ No price data available for zone {zone} in year {year}. Skipping.")
                        print(f"⚠️ No price data available for zone {zone} in year {year}. Skipping.")
                        continue

                    # Calculate average price for the zone
                    df_zone = df_year[zone_nodes]
                    avg_price = df_zone.mean(axis=1).mean()
                    zonal_price_map.loc[zone, str(year)] = round(avg_price, 2)

                except Exception as e:
                    log_messages.append(f"❌ Failed processing zone {zone} in year {year}: {e}")
                    print(f"❌ Failed processing zone {zone} in year {year}: {e}")
                    continue

        except Exception as e:
            log_messages.append(f"❌ Failed processing year {year}: {e}")
            print(f"❌ Failed processing year {year}: {e}")
            continue

    # Join all messages into one string, preserving line breaks
    print("✅ Processing complete!")
    log_text = "\n".join(log_messages)
    return zonal_price_map, log_text

########################################### MAP FUNCTIONS ######################################################

# def nordic_grid_map_fromDB(data, db: Database, time_range, OUTPUT_PATH, version, START, END, exchange_rate_NOK_EUR=11.38):
#     """
#     Generate an interactive map displaying nodal prices and branch utilization.
#
#     This function creates a folium map that visualizes:
#     - Nodes representing average nodal prices.
#     - Branches representing line utilization for both AC and DC connections.
#
#     The generated map is saved as an HTML file for easy visualization.
#
#     Parameters:
#         data (Scenario):
#             The simulation data containing node and branch information.
#         db (Database):
#             Database object used to retrieve average prices, utilization rates, and flow data.
#         time_range (list):
#             List specifying the start and end time steps for the simulation.
#         OUTPUT_PATH (str):
#             Directory path where the HTML map file will be saved.
#         version (str):
#             Version identifier for the map output file.
#         START (dict):
#             Dictionary specifying the start date and time (e.g., {'year': 2023, 'month': 1, 'day': 1, 'hour': 0}).
#         END (dict):
#             Dictionary specifying the end date and time in the same format as START.
#         exchange_rate_NOK_EUR (float, optional):
#             Conversion rate from EUR to NOK for displaying prices in both currencies.
#             Default is 11.38.
#
#     Returns:
#         None: The generated map is saved directly to the specified `OUTPUT_PATH`.
#     """
#     # Assuming data.node.zone is a pandas Series or list
#     # Filter out SINK nodes from data.node
#     data.node = data.node[data.node['zone'] != 'SINK']
#
#     # Get the indices of the remaining nodes
#     valid_node_indices = data.node.id
#
#     # Filter out branches in data.dcbranch connected to SINK nodes
#     data.dcbranch = data.dcbranch[
#         data.dcbranch['node_from'].isin(valid_node_indices) &
#         data.dcbranch['node_to'].isin(valid_node_indices)
#         ]
#
#     # Filter out consumers in data.consumer connected to SINK nodes
#     data.consumer = data.consumer[data.consumer['node'].isin(valid_node_indices)]
#
#     avg_nodal_prices = list(map(float, getAverageNodalPricesFromDB(db, time_range)))
#     avg_area_price = {key: float(value) for key, value in getAreaPricesAverageFromDB(data, db, timeMaxMin=time_range).items()}
#     avg_zone_price = getZonePricesAverageFromDB(data, db, time_range)
#     ac_utilisation = list(map(float, getAverageUtilisationFromDB(data, db, time_range, branchtype="ac")))
#     dc_utilisation = list(map(float, getAverageUtilisationFromDB(data, db, time_range, branchtype="dc")))
#     ac_flows = convert_to_float(getAverageBranchFlowsFromDB(db, time_range, branchtype="ac"))
#     dc_flows = convert_to_float(getAverageBranchFlowsFromDB(db, time_range, branchtype="dc"))
#
#     f = folium.Figure(width=700, height=800)
#     m = folium.Map(location=[data.node["lat"].mean(), data.node["lon"].mean()], zoom_start=4.4)
#     m.add_to(f)
#
#     colormap = cm.LinearColormap(['green', 'yellow', 'red'], vmin=min(avg_nodal_prices), vmax=max(avg_nodal_prices))
#     colormap.caption = 'Nodal Prices'
#     colormap.add_to(m)
#
#     for i, price in enumerate(avg_nodal_prices):
#         add_node_marker(data, i, price, avg_area_price, avg_zone_price, m, colormap,exchange_rate_NOK_EUR)
#
#     line_colormap = cm.LinearColormap(['green', 'yellow', 'red'], vmin=0, vmax=1)
#     line_colormap.caption = 'Branch Utilisation'
#     line_colormap.add_to(m)
#
#     add_branch_lines(data, ac_utilisation, ac_flows, 'AC', m, line_colormap)
#     add_branch_lines(data, dc_utilisation, dc_flows, 'DC', m, line_colormap, dashed=True)
#
#     start_str = f"{START['year']}_{START['month']}_{START['day']}_{START['hour']}"
#     end_str = f"{END['year']}_{END['month']}_{END['day']}_{END['hour']}"
#     output_path = OUTPUT_PATH / f'nordic_grid_map_{version}_{start_str}__to__{end_str}.html'
#     m.save(output_path)
#


#
# def nordic_grid_map_fromDB(data, db: Database, time_range, OUTPUT_PATH, version, START, END, exchange_rate_NOK_EUR=11.38):
#     """
#     Generate an interactive map displaying nodal prices and branch utilization, excluding SINK nodes.
#
#     This function creates a folium map that visualizes:
#     - Nodes representing average nodal prices (excluding nodes in SINK area).
#     - Branches representing line utilization for both AC and DC connections (excluding branches connected to SINK nodes).
#
#     The generated map is saved as an HTML file for easy visualization.
#
#     Parameters:
#         data (Scenario):
#             The simulation data containing node and branch information.
#         db (Database):
#             Database object used to retrieve average prices, utilization rates, and flow data.
#         time_range (list):
#             List specifying the start and end time steps for the simulation.
#         OUTPUT_PATH (str):
#             Directory path where the HTML map file will be saved.
#         version (str):
#             Version identifier for the map output file.
#         START (dict):
#             Dictionary specifying the start date and time (e.g., {'year': 2023, 'month': 1, 'day': 1, 'hour': 0}).
#         END (dict):
#             Dictionary specifying the end date and time in the same format as START.
#         exchange_rate_NOK_EUR (float, optional):
#             Conversion rate from EUR to NOK for displaying prices in both currencies.
#             Default is 11.38.
#
#     Returns:
#         None: The generated map is saved directly to the specified `OUTPUT_PATH`.
#     """
#     # Find indices where zone is SINK
#     sink_index = data.node.zone[data.node.zone == 'SINK'].index.tolist()
#
#     # Filter out SINK nodes from data.node
#     sink_ids = data.node.loc[sink_index, 'id'].to_numpy()
#     valid_dc_branch_mask = ~(
#             data.dcbranch['node_from'].isin(sink_ids) |
#             data.dcbranch['node_to'].isin(sink_ids)
#     )
#     data.node = data.node[data.node['zone'] != 'SINK']
#
#     # Get the indices of the remaining nodes
#     valid_node_indices = data.node.id
#
#     # Filter out branches in data.dcbranch connected to SINK nodes
#     data.dcbranch = data.dcbranch[
#         data.dcbranch['node_from'].isin(valid_node_indices) &
#         data.dcbranch['node_to'].isin(valid_node_indices)
#         ].reset_index(drop=True)
#
#     # Filter out consumers in data.consumer connected to SINK nodes
#     data.consumer = data.consumer[data.consumer['node'].isin(valid_node_indices)]
#
#     # Get data from database
#     avg_nodal_prices = list(map(float, getAverageNodalPricesFromDB(db, time_range)))
#     avg_area_price = {key: float(value) for key, value in getAreaPricesAverageFromDB(data, db, timeMaxMin=time_range).items()}
#     avg_zone_price = getZonePricesAverageFromDB(data, db, time_range)
#     ac_utilisation = list(map(float, getAverageUtilisationFromDB(data, db, time_range, branchtype="ac")))
#     dc_utilisation = list(map(float, getAverageUtilisationFromDB(data, db, time_range, branchtype="dc")))
#     ac_flows = convert_to_float(getAverageBranchFlowsFromDB(db, time_range, branchtype="ac"))
#     dc_flows = convert_to_float(getAverageBranchFlowsFromDB(db, time_range, branchtype="dc"))
#
#     # Filter out SINK nodes from avg_nodal_prices
#     avg_nodal_prices_filtered = [price for i, price in enumerate(avg_nodal_prices) if i not in sink_index]
#
#     # Initialize folium map
#     # Use non-SINK nodes for centering the map
#     non_sink_nodes = data.node[~data.node.index.isin(sink_index)]
#     f = folium.Figure(width=700, height=800)
#     m = folium.Map(location=[non_sink_nodes["lat"].mean(), non_sink_nodes["lon"].mean()], zoom_start=4.4)
#     m.add_to(f)
#
#     # Set up colormap for nodal prices (based on filtered prices)
#     colormap = cm.LinearColormap(['green', 'yellow', 'red'],
#                                vmin=min(avg_nodal_prices_filtered) if avg_nodal_prices_filtered else 0,
#                                vmax=max(avg_nodal_prices_filtered) if avg_nodal_prices_filtered else 1)
#     colormap.caption = 'Nodal Prices'
#     colormap.add_to(m)
#
#     # Add node markers only for non-SINK nodes
#     for i, price in enumerate(avg_nodal_prices_filtered):
#         if i not in sink_index:
#             add_node_marker(data, i, price, avg_area_price, avg_zone_price, m, colormap, exchange_rate_NOK_EUR)
#
#     # Set up colormap for branch utilization
#     line_colormap = cm.LinearColormap(['green', 'yellow', 'red'], vmin=0, vmax=1)
#     line_colormap.caption = 'Branch Utilisation'
#     line_colormap.add_to(m)
#
#     # Filter branches to exclude those connected to SINK nodes
#     try:
#         # Filter AC and DC utilisation and flows
#         # valid_dc_branch_mask = valid_dc_branch_mask.reindex(data.dcbranch.index, fill_value=False)
#         ac_utilisation_filtered = ac_utilisation
#         dc_utilisation_filtered = dc_utilisation
#         ac_flows_filtered = ac_flows
#         dc_flows_filtered = [flow for i, flow in enumerate(dc_flows) if i in data.dcbranch[valid_dc_branch_mask].index]
#     except (AttributeError, KeyError) as e:
#         print(f"Warning: Could not filter branches; branch data unavailable or misconfigured: {e}")
#         ac_utilisation_filtered = ac_utilisation
#         dc_utilisation_filtered = dc_utilisation
#         ac_flows_filtered = ac_flows
#         dc_flows_filtered = dc_flows
#
#     # Add branch lines using filtered data
#     add_branch_lines(data, ac_utilisation_filtered, ac_flows_filtered, 'AC', m, line_colormap)
#     add_branch_lines(data, dc_utilisation_filtered, dc_flows_filtered, 'DC', m, line_colormap, dashed=True)
#
#     # Save the map
#     start_str = f"{START['year']}_{START['month']}_{START['day']}_{START['hour']}"
#     end_str = f"{END['year']}_{END['month']}_{END['day']}_{END['hour']}"
#     output_path = OUTPUT_PATH / f'nordic_grid_map_{version}_{start_str}__to__{end_str}.html'
#     m.save(output_path)
#

def nordic_grid_map_fromDB(data, db: Database, time_range, OUTPUT_PATH, version, START, END, exchange_rate_NOK_EUR=11.38):
    """
    Generate an interactive map displaying nodal prices and branch utilization, excluding SINK nodes.

    This function creates a folium map that visualizes:
    - Nodes representing average nodal prices (excluding nodes in SINK area).
    - Branches representing line utilization for both AC and DC connections (excluding branches connected to SINK nodes).

    The generated map is saved as an HTML file for easy visualization.

    Parameters:
        data (Scenario):
            The simulation data containing node and branch information.
        db (Database):
            Database object used to retrieve average prices, utilization rates, and flow data.
        time_range (list):
            List specifying the start and end time steps for the simulation.
        OUTPUT_PATH (str):
            Directory path where the HTML map file will be saved.
        version (str):
            Version identifier for the map output file.
        START (dict):
            Dictionary specifying the start date and time (e.g., {'year': 2023, 'month': 1, 'day': 1, 'hour': 0}).
        END (dict):
            Dictionary specifying the end date and time in the same format as START.
        exchange_rate_NOK_EUR (float, optional):
            Conversion rate from EUR to NOK for displaying prices in both currencies.
            Default is 11.38.

    Returns:
        None: The generated map is saved directly to the specified `OUTPUT_PATH`.
    """
    # Find indices where zone is SINK
    sink_index = data.node.zone[data.node.zone == 'SINK'].index.tolist()

    # Filter out SINK nodes from data.node
    sink_ids = data.node.loc[sink_index, 'id'].to_numpy()
    data.node = data.node[data.node['zone'] != 'SINK'].reset_index(drop=True)

    # Get the indices of the remaining nodes
    valid_node_indices = data.node['id'].to_numpy()

    # Filter out branches in data.dcbranch connected to SINK nodes
    valid_dc_branch_mask = ~(
        data.dcbranch['node_from'].isin(sink_ids) |
        data.dcbranch['node_to'].isin(sink_ids)
    )
    data.dcbranch = data.dcbranch[valid_dc_branch_mask].reset_index(drop=True)

    # Filter out branches in data.branch (for AC branches) connected to SINK nodes
    try:
        valid_ac_branch_mask = ~(
            data.branch['node_from'].isin(sink_ids) |
            data.branch['node_to'].isin(sink_ids)
        )
        data.branch = data.branch[valid_ac_branch_mask].reset_index(drop=True)
    except (AttributeError, KeyError) as e:
        print(f"Warning: Could not filter AC branches; data.branch unavailable or misconfigured: {e}")
        valid_ac_branch_mask = pd.Series([True] * len(data.branch), index=data.branch.index)

    # Filter out consumers in data.consumer connected to SINK nodes
    data.consumer = data.consumer[data.consumer['node'].isin(valid_node_indices)].reset_index(drop=True)

    # Get data from database
    avg_nodal_prices = list(map(float, getAverageNodalPricesFromDB(db, time_range)))
    avg_area_price = {key: float(value) for key, value in getAreaPricesAverageFromDB(data, db, timeMaxMin=time_range).items()}
    avg_zone_price = getZonePricesAverageFromDB(data, db, time_range)
    ac_utilisation = list(map(float, getAverageUtilisationFromDB(data, db, time_range, branchtype="ac")))
    dc_utilisation = list(map(float, getAverageUtilisationFromDB(data, db, time_range, branchtype="dc")))
    ac_flows = convert_to_float(getAverageBranchFlowsFromDB(db, time_range, branchtype="ac"))
    dc_flows = convert_to_float(getAverageBranchFlowsFromDB(db, time_range, branchtype="dc"))

    # Filter out SINK nodes from avg_nodal_prices
    avg_nodal_prices_filtered = [price for i, price in enumerate(avg_nodal_prices) if i not in sink_index]

    # Initialize folium map
    f = folium.Figure(width=700, height=800)
    m = folium.Map(location=[data.node["lat"].mean(), data.node["lon"].mean()], zoom_start=4.4)
    m.add_to(f)

    # Set up colormap for nodal prices (based on filtered prices)
    colormap = cm.LinearColormap(
        ['green', 'yellow', 'red'],
        vmin=min(avg_nodal_prices_filtered) if avg_nodal_prices_filtered else 0,
        vmax=max(avg_nodal_prices_filtered) if avg_nodal_prices_filtered else 1
    )
    colormap.caption = 'Nodal Prices'
    colormap.add_to(m)

    # Add node markers for non-SINK nodes
    valid_node_indices_list = data.node.index.tolist()  # Indices of non-SINK nodes
    for i, idx in enumerate(valid_node_indices_list):
        price = avg_nodal_prices_filtered[i]
        add_node_marker(data, idx, price, avg_area_price, avg_zone_price, m, colormap, exchange_rate_NOK_EUR)

    # Set up colormap for branch utilization
    line_colormap = cm.LinearColormap(['green', 'yellow', 'red'], vmin=0, vmax=1)
    line_colormap.caption = 'Branch Utilisation'
    line_colormap.add_to(m)

    # Filter utilisation and flows for branches
    try:
        # Filter AC utilisation and flows
        ac_utilisation_filtered = [util for i, util in enumerate(ac_utilisation) if valid_ac_branch_mask.iloc[i]]
        ac_flows_filtered = [flow for i, flow in enumerate(ac_flows) if valid_ac_branch_mask.iloc[i]]
        # Filter DC utilisation and flows
        dc_utilisation_filtered = dc_utilisation # [util for i, util in enumerate(dc_utilisation) if valid_dc_branch_mask.iloc[i]]
        # dc_flows_filtered = [flow for i, flow in enumerate(dc_flows) if valid_dc_branch_mask.iloc[i]]
        dc_flows_filtered = [
            [flow for i, flow in enumerate(flow_sublist) if valid_dc_branch_mask.iloc[i]]
            for flow_sublist in dc_flows
        ]
    except (AttributeError, KeyError, IndexError) as e:
        print(f"Warning: Could not filter branches; branch data unavailable or misconfigured: {e}")
        ac_utilisation_filtered = ac_utilisation
        dc_utilisation_filtered = dc_utilisation
        ac_flows_filtered = ac_flows
        dc_flows_filtered = dc_flows

    # Add branch lines using filtered data
    add_branch_lines(data, ac_utilisation_filtered, ac_flows_filtered, 'AC', m, line_colormap)
    add_branch_lines(data, dc_utilisation_filtered, dc_flows_filtered, 'DC', m, line_colormap, dashed=True)

    # Save the map
    start_str = f"{START['year']}_{START['month']}_{START['day']}_{START['hour']}"
    end_str = f"{END['year']}_{END['month']}_{END['day']}_{END['hour']}"
    output_path = OUTPUT_PATH / f'nordic_grid_map_{version}_{start_str}__to__{end_str}.html'
    m.save(output_path)



def nordic_grid_map(data, res, time_max_min, OUTPUT_PATH, version, exchange_rate_NOK_EUR=11.38):
    """
    Generate an interactive map of the Nordic power grid showing nodal prices, zonal prices, and line utilization.

    This function creates a folium-based map that visualizes:
    - Average nodal electricity prices using a color-coded scale.
    - Average zonal prices representing regional electricity markets.
    - National average prices (area prices) for each country.
    - Transmission line utilization for both AC and DC branches, along with flow data.

    The map offers a spatial and intuitive overview of price levels and network loading
    over a specified time period. It is saved as an HTML file for interactive exploration.

    Parameters:
        data (Scenario):
            The scenario data including node locations and network topology.
        res (ResultHandler):
            Object providing access to simulation results such as prices, utilization, and flows.
        time_max_min (list):
            A list indicating the time range to average over.
        OUTPUT_PATH (str or Path):
            Directory where the HTML file will be stored.
        version (str):
            Identifier to distinguish map output versions.
        exchange_rate_NOK_EUR (float, optional):
            Exchange rate for converting EUR to NOK when displaying prices. Default is 11.38.

    Returns:
        None: The HTML map is saved to `OUTPUT_PATH`.
    """

    avg_nodal_prices = list(map(float, res.getAverageNodalPrices(time_max_min)))
    avg_zonal_prices = res.getAverageZonalPrices(time_max_min)
    avg_area_price = {key: float(value) for key, value in res.getAreaPricesAverage(timeMaxMin=time_max_min).items()}
    ac_utilisation = list(map(float, res.getAverageUtilisation(time_max_min, branchtype="ac")))
    ac_flows = convert_to_float(res.getAverageBranchFlows(time_max_min, branchtype="ac"))
    dc_utilisation = list(map(float, res.getAverageUtilisation(time_max_min, branchtype="dc")))
    dc_flows = convert_to_float(res.getAverageBranchFlows(time_max_min, branchtype="dc"))

    f = folium.Figure(width=700, height=800)
    m = folium.Map(location=[data.node["lat"].mean(), data.node["lon"].mean()], zoom_start=4.4)
    m.add_to(f)

    colormap = cm.LinearColormap(['green', 'yellow', 'red'], vmin=min(avg_nodal_prices), vmax=max(avg_nodal_prices))
    colormap.caption = 'Nodal Prices'
    colormap.add_to(m)

    for i, price in enumerate(avg_nodal_prices):
        add_node_marker(data, i, price, avg_area_price, avg_zonal_prices, m, colormap, exchange_rate_NOK_EUR)


    line_colormap = cm.LinearColormap(['green', 'yellow', 'red'], vmin=0, vmax=1)
    line_colormap.caption = 'Branch Utilisation'
    line_colormap.add_to(m)

    add_branch_lines(data, ac_utilisation, ac_flows, 'AC', m, line_colormap)
    add_branch_lines(data, dc_utilisation, dc_flows, 'DC', m, line_colormap, dashed=True)

    output_path = OUTPUT_PATH / f'nordic_grid_map_{version}.html'
    m.save(output_path)



"""Flow Based Functions"""
def create_price_and_utilization_map(data, res, time_max_min, output_path, dc=None):
    """
        Generate a folium map displaying nodal prices and branch utilization.

        This function creates a map with nodes representing average prices and branches representing
        line utilization for the given time range. The map is saved as an HTML file.

        Parameters:
            data (Scenario):        Simulation data containing nodes and branches.
            res (Result):           Result object with nodal prices, utilization, and flows.
            time_max_min (list):    List specifying the start and end time steps for the simulation.
            output_path (str):      Path where the HTML map file will be saved.

        Returns:
            None
    """

    nodal_prices = list(map(float, res.getAverageNodalPrices(time_max_min)))
    avg_area_price = {key: float(value) for key, value in res.getAreaPricesAverage(timeMaxMin=time_max_min).items()}
    ac_utilisation = list(map(float, res.getAverageUtilisation(time_max_min, branchtype="ac")))
    ac_flows = convert_to_float(res.getAverageBranchFlows(time_max_min, branchtype="ac"))
    if dc is not None:
        dc_utilisation = list(map(float, res.getAverageUtilisation(time_max_min, branchtype="dc")))
        dc_flows = convert_to_float(res.getAverageBranchFlows(time_max_min, branchtype="dc"))

    f = folium.Figure(width=700, height=800)
    m = folium.Map(location=[data.node["lat"].mean(), data.node["lon"].mean()], zoom_start=4.4)
    m.add_to(f)

    colormap = cm.LinearColormap(['green', 'yellow', 'red'], vmin=min(nodal_prices), vmax=max(nodal_prices))
    colormap.caption = 'Nodal Prices'
    colormap.add_to(m)

    for i, price in enumerate(nodal_prices):
        add_node_marker(data, i, price, avg_area_price, m, colormap)

    line_colormap = cm.LinearColormap(['green', 'yellow', 'red'], vmin=0, vmax=1)
    line_colormap.caption = 'Branch Utilisation'
    line_colormap.add_to(m)

    add_branch_lines(data, ac_utilisation, ac_flows[2], 'AC', m, line_colormap)
    if dc is not None:
        add_branch_lines(data, dc_utilisation, dc_flows[2], 'DC', m, line_colormap, dashed=True)

    m.save(output_path)
    display(m)

def convert_to_float(data):
    """Recursively convert np.float64 values in lists to Python float"""
    if isinstance(data, list):
        return [convert_to_float(item) for item in data]
    elif isinstance(data, np.float64):
        return float(data)
    else:
        return data

def create_price_and_utilization_map_FromDB(data: GridData, db: Database, time_max_min, output_path):
    """
    Generate a folium map displaying nodal prices and branch utilization.

    This function creates a map with nodes representing average prices and branches representing
    line utilization for the given time range. The map is saved as an HTML file.

    Parameters:
        data (Scenario):        Simulation data containing nodes and branches.
        db (Database):          Result object with nodal prices, utilization, and flows.
        time_max_min (list):    List specifying the start and end time steps for the simulation.
        output_path (str):      Path where the HTML map file will be saved.
        eur_to_nok (float):     Conversion rate from EUR to NOK for price display.

    Returns:
        None
    """
    nodal_prices = list(map(float, getAverageNodalPricesFromDB(db, time_max_min)))
    avg_area_price = {key: float(value) for key, value in getAreaPricesAverageFromDB(data, db, timeMaxMin=time_max_min).items()}
    ac_utilisation = list(map(float, getAverageUtilisationFromDB(data, db, time_max_min, branchtype="ac")))
    dc_utilisation = list(map(float, getAverageUtilisationFromDB(data, db, time_max_min, branchtype="dc")))
    ac_flows = convert_to_float(getAverageBranchFlowsFromDB(db, time_max_min, branchtype="ac"))
    dc_flows = convert_to_float(getAverageBranchFlowsFromDB(db, time_max_min, branchtype="dc"))

    f = folium.Figure(width=700, height=800)
    m = folium.Map(location=[data.node["lat"].mean(), data.node["lon"].mean()], zoom_start=4.4)
    m.add_to(f)

    colormap = cm.LinearColormap(['green', 'yellow', 'red'], vmin=min(nodal_prices), vmax=max(nodal_prices))
    colormap.caption = 'Nodal Prices'
    colormap.add_to(m)

    for i, price in enumerate(nodal_prices):
        add_node_marker(data, i, price, avg_area_price, m, colormap)

    line_colormap = cm.LinearColormap(['green', 'yellow', 'red'], vmin=0, vmax=1)
    line_colormap.caption = 'Branch Utilisation'
    line_colormap.add_to(m)

    add_branch_lines(data, ac_utilisation, ac_flows[2], 'AC', m, line_colormap)
    add_branch_lines(data, dc_utilisation, dc_flows[2], 'DC', m, line_colormap, dashed=True)

    m.save(output_path)
    display(m)


def add_node_marker(data, index, price, avg_national_prices, avg_zonal_price, m, colormap, exchange_rate_NOK_EUR):
    """
    Add a marker to the map for a specific node, displaying its price and zone information.

    Parameters:
        data (Scenario):            Simulation data containing node information.
        index (int):                Node index within the data.
        price (float):              Nodal price for the given node.
        avg_national_prices (dict): Dictionary of average national prices by area.
        m (folium.Map):             Folium map object to which the marker will be added.
        colormap (LinearColormap):  Colormap for representing nodal prices.
        eur_to_nok (float):         Conversion rate from EUR to NOK. - NOT INCLUDED ANYMORE

    Returns:
        None
    """
    lat = data.node.loc[index, 'lat']
    lon = data.node.loc[index, 'lon']
    node_idx = data.node.loc[index, 'index']
    node_id = data.node.loc[index, 'id']
    node_zone = data.node.loc[index, 'zone']
    area = data.node.loc[index, 'area']
    zonal_price = avg_zonal_price.get(node_zone, 'N/A')
    area_price = avg_national_prices.get(area, 'N/A')
    nodal_price_nok = (price * exchange_rate_NOK_EUR) / 10
    zonal_price_nok = (zonal_price * exchange_rate_NOK_EUR) / 10
    area_price_nok = (area_price * exchange_rate_NOK_EUR) / 10

    # EUR_TO_ORE_PER_KWH = eur_to_nok / 1000 * 100
    # price_nok = price * EUR_TO_ORE_PER_KWH
    # area_price_nok = area_price * EUR_TO_ORE_PER_KWH if isinstance(area_price, (int, float)) else 'N/A'

    popup = folium.Popup(
        f"<b>Node index:</b> {node_idx}<br>"
        f"<b>Node id:</b> {node_id}<br>"
        f"<b>Zone:</b> {node_zone}<br>"
        f"<b>Price:</b> {price:.2f} EUR/MWh = {nodal_price_nok:.2f} Øre/KWh <br>"
        f"<b>Zonal Price:</b> {f'{zonal_price:.2f} EUR/MWh = {zonal_price_nok:.2f} Øre/KWh' if isinstance(zonal_price, (int, float)) else 'N/A'}<br>"
        f"<b>National Price:</b> {f'{area_price:.2f} EUR/MWh = {area_price_nok:.2f} Øre/KWh' if isinstance(area_price_nok, (int, float)) else 'N/A'}<br>",
        max_width=300
    )
    folium.CircleMarker(
        location=[lat, lon],
        radius=12,
        popup=popup,
        color=colormap(price),
        fill=True,
        fill_color=colormap(price),
        fill_opacity=0.7
    ).add_to(m)



def add_branch_lines(data, utilisation, flows, branch_type, m, line_colormap, dashed=False):
    """
    Add lines representing AC or DC branches on the map, showing flow and utilization.

    Parameters:
        data (Scenario):                Simulation data containing branch information.
        utilisation (list):             List of average utilization values for each branch.
        flows (list):                   List of average flow values for each branch.
        branch_type (str):              Type of branch ("AC" or "DC").
        m (folium.Map):                 Folium map object to which the lines will be added.
        line_colormap (LinearColormap): Colormap for representing utilization.
        dashed (bool):                  If True, displays lines as dashed; used for DC branches.

    Returns:
        None
    """
    branches = data.branch if branch_type == 'AC' else data.dcbranch
    for idx, row in branches.iterrows():
        utilisation_percent = utilisation[idx] * 100
        nodeA = data.node.loc[data.node['id'] == row['node_from']].iloc[0]
        nodeB = data.node.loc[data.node['id'] == row['node_to']].iloc[0]
        line_color = line_colormap(utilisation[idx])

        popup_content = folium.Popup(
            f"<b>{branch_type} Line</b><br>"
            f"<b>Power Flow:</b> {flows[2][idx]:.2f} MW<br>"
            f"<b>Utilisation:</b> {utilisation_percent:.2f}%",
            max_width=150
        )

        folium.PolyLine(
            locations=[(nodeA['lat'], nodeA['lon']), (nodeB['lat'], nodeB['lon'])],
            color=line_color,
            weight=5,
            dash_array="5, 10" if dashed else None,
            opacity=0.7,
            popup=popup_content
        ).add_to(m)

        zoneA = row['node_from'][:3]
        zoneB = row['node_to'][:3]

        if zoneA != zoneB:
            # Convert endpoints to Mercator
            Ax, Ay = to_web_mercator(nodeA['lat'], nodeA['lon'])
            Bx, By = to_web_mercator(nodeB['lat'], nodeB['lon'])

            # Always compute the midpoint as the average
            Mx = (Ax + Bx) / 2
            My = (Ay + By) / 2

            # Determine the direction based on flows
            if flows[0][idx] >= flows[1][idx]:
                dx = Bx - Ax
                dy = By - Ay
            else:
                dx = Ax - Bx
                dy = Ay - By

            angle = math.degrees(math.atan2(dx, dy)) % 360
            # print(f"From: {row['node_from']} → {row['node_to']}, dx={dx:.1f}, dy={dy:.1f}, angle={angle:.1f}, rotation={angle:.1f}")
            mid_lat, mid_lon = from_web_mercator(Mx, My)

            folium.Marker(
                location=[mid_lat, mid_lon],
                icon=folium.DivIcon(
                    html=svg_arrow_icon(angle),
                    icon_size=(24,24),
                    icon_anchor=(12,12)
                ),
                popup="Flow direction"
            ).add_to(m)


def to_web_mercator(lat, lon):
    R = 6378137.0  # Earth radius
    x = R * math.radians(lon)
    y = R * math.log(math.tan(math.pi/4 + math.radians(lat)/2))
    return x, y

def from_web_mercator(x, y):
    R = 6378137.0
    lon = math.degrees(x / R)
    lat = math.degrees(2 * math.atan(math.exp(y / R)) - math.pi/2)
    return lat, lon


def svg_arrow_icon(angle: float, color="black") -> str:
    return f"""
    <div style="transform: rotate({angle}deg); width: 24px; height: 24px;">
        <svg width="24" height="24" viewBox="0 0 24 24" 
             xmlns="http://www.w3.org/2000/svg" 
             style="transform: rotate(0deg);">
            <polygon points="12,0 4,20 12,16 20,20" fill="{color}" />
        </svg>
    </div>
    """

  

def get_interconnections(data: GridData):
    """
    Retrieve and filter cross-country AC and DC branch data from CSV files.

    This function reads AC and DC branch data from CSV files, filters for
    cross-country connections, and returns separate dictionaries for each.

    Parameters:
        datapath_GridData (Path): Path to the grid data directory.

    Returns:
        tuple: Two dictionaries (AC_dict, DC_dict), where each dictionary has branch indices as keys
               and tuples of (node_from, node_to) as values for AC and DC branches, respectively.
    """
    AC_cross_country_connections, DC_cross_country_connections = filter_cross_country_connections(data)

    # Create dictionaries with index as key and tuple (node_from, node_to) as value for each
    AC_cross_country_dict = {
        int(row['Unnamed: 0']): (row['node_from'], row['node_to'])
        for _, row in AC_cross_country_connections.iterrows()
    }
    DC_cross_country_dict = {
        int(row['Unnamed: 0']): (row['node_from'], row['node_to'])
        for _, row in DC_cross_country_connections.iterrows()
    }
    return AC_cross_country_dict, DC_cross_country_dict



def calculate_interconnections_flow(db, datapath_GridData, time_max_min):
    """
    Calculate import, export, and average absolute flow for AC and DC interconnections.

    This function retrieves flow data for all cross-country AC and DC branches over the
    simulation period, calculates the total and average flow, and returns the results as a DataFrame.

    Parameters:
        db (Database):              Database instance containing branch flow data.
        datapath_GridData (Path):   Path to the grid data directory for retrieving branch information.
        time_max_min (list):        List containing the start and end time steps for the simulation.

    Returns:
        pd.DataFrame: DataFrame with columns [index, type, from, to, import [MWh], export [MWh],
                      average_absolute_flow [MW]], containing flow data for each interconnection.
    """
    # Get cross-country interconnections
    AC_cross_country_dict, DC_cross_country_dict = get_interconnections(datapath_GridData)

    # Combine AC and DC connections into a single dictionary with branch type for easier processing
    all_interconnections = {
        **{(idx, 'AC'): nodes for idx, nodes in AC_cross_country_dict.items()},
        **{(idx, 'DC'): nodes for idx, nodes in DC_cross_country_dict.items()}
    }
    flow_data = []
    num_time_steps = time_max_min[1]
    # Loop through each connection
    for (branch_index, branch_type), (node_from, node_to) in all_interconnections.items():
        # Get flow data for the branch
        all_branch_flows = db.getResultBranchFlow(branch_index, time_max_min, ac=(branch_type == 'AC'))

        absolute_flows = [abs(flow) for flow in all_branch_flows]
        total_absolute_flow = sum(absolute_flows)
        average_absolute_flow = total_absolute_flow / num_time_steps  # Average absolute flow

        # Calculate import and export flow based on flow direction
        import_flow = sum(flow for flow in all_branch_flows if flow < 0)  # negative flow = import
        export_flow = sum(flow for flow in all_branch_flows if flow > 0)  # positive flow = export

        # Append results to flow data list
        flow_data.append({
            'index': branch_index,
            'type': branch_type,
            'from': node_from,
            'to': node_to,
            'import [MWh]': abs(import_flow),
            'export [MWh]': abs(export_flow),
            'average_absolute_flow [MW]': average_absolute_flow
        })
    # Create DataFrame from the collected flow data
    flow_df = pd.DataFrame(flow_data, columns=['index', 'type', 'from', 'to', 'import [MWh]', 'export [MWh]',
                                               'average_absolute_flow [MW]'])
    return flow_df


def filter_cross_country_connections(data: GridData):
    """
    Filter cross-country connections between AC and DC branches.
    """
    AC_branch_df = data.branch
    DC_branch_df = data.dcbranch
    AC_cross_country_connections = AC_branch_df[AC_branch_df['node_from'].str[:2] != AC_branch_df['node_to'].str[:2]]
    DC_cross_country_connections = DC_branch_df[DC_branch_df['node_from'].str[:2] != DC_branch_df['node_to'].str[:2]]
    return AC_cross_country_connections, DC_cross_country_connections


def filter_cross_border_connections(data: GridData):
    """
    Filter cross-border connections between zones and countries, exclude connections within same zone.
    """
    AC_branch_df = data.branch
    DC_branch_df = data.dcbranch
    AC_cross_border_connections = AC_branch_df[AC_branch_df['node_from'].str[:3] != AC_branch_df['node_to'].str[:3]]
    DC_cross_border_connections = DC_branch_df[DC_branch_df['node_from'].str[:3] != DC_branch_df['node_to'].str[:3]]
    return AC_cross_border_connections, DC_cross_border_connections



def get_connections(data: GridData, chosen_connections):
    """
    Retrieve and filter AC and DC branch data from CSV files.

    This function reads AC and DC branch data from CSV files, filters for
    cross-country connections, and returns separate dictionaries for each.

    Parameters:
        datapath_GridData (Path): Path to the grid data directory.

    Returns:
        tuple: Two dictionaries (AC_dict, DC_dict), where each dictionary has branch indices as keys
               and tuples of (node_from, node_to) as values for AC and DC branches, respectively.
    """
    AC_connections, DC_connections = filter_connections_by_list(data, chosen_connections)

    # Create dictionaries with index as key and tuple (node_from, node_to) as value for each
    AC_dict = {
        int(row['Unnamed: 0']): (row['node_from'], row['node_to'])
        for _, row in AC_connections.iterrows()
    }
    DC_dict = {
        int(row['Unnamed: 0']): (row['node_from'], row['node_to'])
        for _, row in DC_connections.iterrows()
    }
    return AC_dict, DC_dict



def filter_connections_by_list(data: GridData, chosen_connections=None):
    AC_branch_df = data.branch
    DC_branch_df = data.dcbranch

    if chosen_connections:
        # Filter AC branches where the pair [node_from, node_to] matches any in chosen_connections
        AC_connections = AC_branch_df[
            AC_branch_df.apply(lambda row: [row['node_from'], row['node_to']] in chosen_connections
                                           or [row['node_to'], row['node_from']] in chosen_connections, axis=1)
        ]
        # Example filter for DC branches: Difference between node_from and node_to countries
        DC_connections = DC_branch_df[
            DC_branch_df.apply(lambda row: [row['node_from'], row['node_to']] in chosen_connections
                                           or [row['node_to'], row['node_from']] in chosen_connections, axis=1)
        ]
    else:
        # If no chosen connections are provided, return all branches
        AC_connections = AC_branch_df
        DC_connections = DC_branch_df

    return AC_connections, DC_connections



def plot_LDC_interconnections(data, db, grid_data_path, time_max_min, OUTPUT_PATH_PLOTS, tex_font):
    AC_interconnections, DC_interconnections = filter_cross_country_connections(data)

    AC_interconnections_capacity = AC_interconnections['capacity']
    DC_interconnections_capacity = DC_interconnections['capacity']

    # Get cross-country interconnections
    AC_cross_country_dict, DC_cross_country_dict = get_interconnections(grid_data_path)

    # AC
    flow_data_AC = []
    branch_type = "AC"
    # Loop through each connection
    for branch_index, (node_from, node_to) in AC_cross_country_dict.items():
        # Get flow data for the branch
        branch_flows = db.getResultBranchFlow(branch_index, time_max_min, ac=True)
        absolute_flows = [abs(flow) for flow in branch_flows]
        max_capacity = AC_interconnections_capacity[branch_index]

        # Append results to flow data list
        flow_data_AC.append({
            'index': branch_index,
            'type': branch_type,
            'from': node_from,
            'to': node_to,
            'load [MW]': absolute_flows,
            'capacity [MW]': max_capacity,
        })
    for item in flow_data_AC:
        item['load [MW]'] = sorted(item['load [MW]'], reverse=True)


    # DC
    flow_data_DC = []
    branch_type = "DC"
    # Loop through each connection
    for branch_index, (node_from, node_to) in DC_cross_country_dict.items():
        # Get flow data for the branch
        branch_flows = db.getResultBranchFlow(branch_index, time_max_min, ac=False)
        absolute_flows = [abs(flow) for flow in branch_flows]
        max_capacity = DC_interconnections_capacity[branch_index]

        # Append results to flow data list
        flow_data_DC.append({
            'index': branch_index,
            'type': branch_type,
            'from': node_from,
            'to': node_to,
            'load [MW]': absolute_flows,
            'capacity [MW]': max_capacity,
        })
    for item in flow_data_DC:
        item['load [MW]'] = sorted(item['load [MW]'], reverse=True)

    flow_df = pd.concat([
        pd.DataFrame(flow_data_AC),
        pd.DataFrame(flow_data_DC)
    ], ignore_index=True)

    # Plot load duration curves for each interconnection
    for index, row in flow_df.iterrows():
        plt.figure(figsize=(10, 6))  # Opprett en ny figur for hver interconnection
        plt.plot(row['load [MW]'], label=f"From: {row['from']} To: {row['to']}")
        plt.axhline(y=row['capacity [MW]'], color='red', linestyle='--', label=f"Maximum capacity: {row['capacity [MW]']:.2f} MW")
        max_y = row['capacity [MW]']
        plt.ylim(0, max_y + 50)
        plt.title(f"Load duration curve for {row['type']} connection {row['from']} <--> {row['to']}")
        plt.xlabel("Time [hours]")
        plt.ylabel("Load [MW]")
        plt.legend(loc='upper right')
        plt.grid(True)
        plt.tight_layout()
        if tex_font:
            plt.rcParams.update({
                "text.usetex": True,
                "font.family": "serif",
                "font.serif": ["Computer Modern Roman"]})
        plot_file_name = OUTPUT_PATH_PLOTS / f"load_duration_curve_{row['from']}_{row['to']}.pdf"
        plt.savefig(plot_file_name)  # Save the plot as a PDF file
        plt.close()
        # plt.show()
    return flow_df




def getEnergyBalanceZoneLevel(all_nodes, totalDemand, totalProduction, totalLoadShedding, flow_data, OUTPUT_PATH, VERSION, START):
    """

    :param all_nodes:
    :param totalDemand:
    :param totalProduction:
    :param totalLoadShedding:
    :param flow_data:
    :param OUTPUT_PATH:
    :return: zone_energyBalance
    """
    # Extract node names from all_nodes (list of tuples)
    node_names = [node for node in all_nodes]

    # Create node-to-zone mapping
    node_to_zone = {}
    for node in node_names:
        if '_' in node:
            # Take prefix before first underscore (e.g., 'DK1_3' -> 'DK1', 'SE3_hub_east' -> 'SE3')
            zone = node.split('_')[0]
        else:
            # No underscore (e.g., 'GB', 'DE') -> use full name as zone
            zone = node
        node_to_zone[node] = zone

    # Get all unique zones
    all_zones = sorted(set(node_to_zone.values()))

    # Aggregate node-level data by zone
    zone_demand = {zone: 0 for zone in all_zones}
    zone_production = {zone: 0 for zone in all_zones}
    zone_load_shedding = {zone: 0 for zone in all_zones}

    # Sum Demand and Production for each zone
    for node in node_names:
        zone = node_to_zone[node]
        zone_demand[zone] += totalDemand.get(node, 0)
        zone_production[zone] += totalProduction.get(node, 0)


    # Sum Load Shedding for each zone (totalLoadShedding is a list aligned with all_nodes)
    for i, node in enumerate(all_nodes):
        zone = node_to_zone[node]
        if i < len(totalLoadShedding):
            zone_load_shedding[zone] += totalLoadShedding[i]

    # Assuming flow_data is a list of lists or a DataFrame
    # Convert to DataFrame for easier handling if not already
    if isinstance(flow_data, list):
        flow_df = pd.DataFrame(flow_data[1:], columns=flow_data[0])
    else:
        flow_df = flow_data

    # Initialize dictionaries for zone-level import and export
    zone_imports = {zone: 0 for zone in all_zones}
    zone_exports = {zone: 0 for zone in all_zones}


    import ast
    # Process each line in flow_data
    for _, row in flow_df.iterrows():
        from_node = row['from']
        to_node = row['to']
        loads = row['load [MW]']  # List of load values

        # Ensure loads is a list or array
        if isinstance(loads, str):
            loads = ast.literal_eval(loads)  # Safer parsing
        loads = np.array(loads)

        # Get zones for from and to nodes
        from_zone = node_to_zone.get(from_node)
        to_zone = node_to_zone.get(to_node)

        # Skip if nodes are in the same zone or if zones are not defined
        if from_zone is None or to_zone is None or from_zone == to_zone:
            continue

        # Positive load: flow from 'from' to 'to'
        # - 'from' zone exports (positive load)
        # - 'to' zone imports (positive load)
        positive_loads = loads[loads > 0]
        if len(positive_loads) > 0:
            zone_exports[from_zone] += sum(positive_loads)
            zone_imports[to_zone] += sum(positive_loads)

        # Negative load: flow from 'to' to 'from'
        # - 'to' zone exports (absolute of negative load)
        # - 'from' zone imports (absolute of negative load)
        negative_loads = loads[loads < 0]
        if len(negative_loads) > 0:
            zone_exports[to_zone] += sum(-negative_loads)  # Absolute value
            zone_imports[from_zone] += sum(-negative_loads)  # Absolute value


    # Prepare zone-level EBData
    EBDataZoneLevel = {
        'Zone': all_zones,
        'Demand': [zone_demand[zone] for zone in all_zones],
        'Production': [zone_production[zone] for zone in all_zones],
        'Load Shedding': [zone_load_shedding[zone] for zone in all_zones],
        'Import': [zone_imports[zone] for zone in all_zones],
        'Export': [zone_exports[zone] for zone in all_zones],
    }

    # Create the zone-level energyBalance DataFrame
    zone_energyBalance = pd.DataFrame(EBDataZoneLevel)
    zone_energyBalance['Balance_mls'] = zone_energyBalance['Production'] - zone_energyBalance['Demand'] + zone_energyBalance['Load Shedding']
    zone_energyBalance['Balance_uls'] = zone_energyBalance['Production'] - zone_energyBalance['Demand']
    zone_energyBalance['NetExport'] = zone_energyBalance['Export'] - zone_energyBalance['Import']
    zone_energyBalance.to_csv(OUTPUT_PATH / f'zone_energy_balance_{VERSION}_{START['year']}.csv', index=False)
    return zone_energyBalance


# Initialize dictionaries to store import and export sums for each node
def getEnergyBalanceNodeLevel(all_nodes, totalDemand, totalProduction, totalLoadShedding, flow_data, OUTPUT_PATH, VERSION, START):
    """
    :param all_nodes:
    :param totalDemand:
    :param totalProduction:
    :param totalLoadShedding:
    :param flow_data:
    :param OUTPUT_PATH:
    :return: energyBalance
    """
    node_imports = {node: 0 for node in all_nodes}
    node_exports = {node: 0 for node in all_nodes}

    # Process each line in flow_data
    for _, row in flow_data.iterrows():
        from_node = row['from']
        to_node = row['to']
        loads = row['load [MW]']  # List of load values

        # Ensure loads is a list or array
        if isinstance(loads, str):
            loads = eval(loads)  # Convert string representation to list if needed
        loads = np.array(loads)

        # Positive load: flow from 'from' to 'to'
        # - 'from' node exports (positive load)
        # - 'to' node imports (positive load)
        positive_loads = loads[loads > 0]
        if len(positive_loads) > 0:
            if from_node in node_exports:
                node_exports[from_node] += sum(positive_loads)
            if to_node in node_imports:
                node_imports[to_node] += sum(positive_loads)

        # Negative load: flow from 'to' to 'from'
        # - 'to' node exports (absolute of negative load)
        # - 'from' node imports (absolute of negative load)
        negative_loads = loads[loads < 0]
        if len(negative_loads) > 0:
            if to_node in node_exports:
                node_exports[to_node] += sum(-negative_loads)  # Absolute value
            if from_node in node_imports:
                node_imports[from_node] += sum(-negative_loads)  # Absolute value

    # Prepare EBData with all components, including Balance
    # Handle totalLoadShedding as a list aligned with all_nodes
    load_shedding_values = [totalLoadShedding[i] if i < len(totalLoadShedding) else 0 for i in range(len(all_nodes))]

    EBData = {
        'Node': all_nodes,
        'Demand': [totalDemand.get(n, 0) for n in all_nodes],
        'Load Shedding': load_shedding_values,
        'Production': [totalProduction.get(n, 0) for n in all_nodes],
        'Import': [node_imports.get(n, 0) for n in all_nodes],
        'Export': [node_exports.get(n, 0) for n in all_nodes],
    }

    # Create the energyBalance DataFrame
    energyBalance = pd.DataFrame(EBData)
    energyBalance['Balance_mls'] = energyBalance['Production'] - energyBalance['Demand'] + energyBalance['Load Shedding']
    energyBalance['Balance_uls'] = energyBalance['Production'] - energyBalance['Demand']
    energyBalance['NetExport'] = energyBalance['Export'] - energyBalance['Import']
    # Save the DataFrame to a CSV file for reference
    energyBalance.to_csv(OUTPUT_PATH / f'node_energy_balance_{VERSION}_{START['year']}.csv', index=False)
    return energyBalance






################################### ANALYSIS FUNCTIONS #############################################################


# Time handling function using Python's built-in datetime objects.
def get_hour_range(YEAR_START, YEAR_END, TIMEZONE, start, end):
    """
    Beregner timeindeks for et gitt tidsintervall basert på simuleringens starttidspunkt.
    Simuleringen starter ved YEAR_START (Gitt i "General Configurations" ).
    Tar inn to tidspunkt (start og slutt) i dictionary-format og returnerer
    indeksene for disse tidspunktene i forhold til simuleringens start.

    Eksempel:
    YEAR_START: Startåret for simuleringsperioden (SQL)
    YEAR_END: Sluttåret for simuleringsperioden (SQL)
    TIMEZONE = ZoneInfo("UTC")
    START = {"year": 2005, "month": 5, "day": 5, "hour": 14}
    END = {"year": 2005, "month": 5, "day": 8, "hour": 14}
    get_hour_range(START, END)

    Funksjonen håndterer skuddår.
    Gir feilmelding dersom start og slutt ikke er innenfor perioden som SQL-filen har simulert over.
    """

    if not (YEAR_START <= start["year"] <= YEAR_END) or not (YEAR_START <= end["year"] <= YEAR_END):
        raise ValueError(f"Input years must be within {YEAR_START}-{YEAR_END}")

    start_time = datetime(YEAR_START, 1, 1, 0, 0, tzinfo=TIMEZONE)

    # Definer start- og sluttidspunktet basert på input
    start_datetime = datetime(start["year"], start["month"], start["day"], start["hour"], 0, tzinfo=TIMEZONE)
    end_datetime = datetime(end["year"], end["month"], end["day"], end["hour"], 0, tzinfo=TIMEZONE)

    # Beregn timeindeks
    start_hour_index = int((start_datetime - start_time).total_seconds() / 3600)
    end_hour_index = int((end_datetime - start_time).total_seconds() / 3600) + 1

    print(f"Start hour index: {start_hour_index}")
    print(f"End hour index: {end_hour_index}")
    print(f"Time steps: {end_hour_index - start_hour_index}")

    return start_hour_index, end_hour_index


def auto_adjust_column_width(ws, max_width=30):
    for col_idx, col_cells in enumerate(ws.columns, 1):  # 1-based index
        lengths = []

        for cell in col_cells:
            if cell.value is not None:
                value_length = len(str(cell.value))
                lengths.append(value_length)

        if lengths:
            # Take 90th percentile length or max header length, whichever is greater
            content_width = int(np.percentile(lengths, 90))
            header_width = lengths[0]  # First cell is header
            adjusted_width = min(max(content_width, header_width) + 2, max_width)

            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = adjusted_width


#### Get production in specific node

def GetProductionAtSpecificNodes(Nodes, data: GridData, database: Database, start_hour, end_hour):
    print(f'🔄 Samler produksjon fra {len(Nodes)} noder...')

    # === 1. Lag mappings og samle alle generatorer ===
    node_idx = data.node[data.node['id'].isin(Nodes)].index.tolist()
    gen_idx = [[gen for gen in data.getGeneratorsAtNode(idx)] for idx in node_idx]
    gen_type = [[data.generator.loc[gen, "type"] for gen in gens] for gens in gen_idx]

    # Flat generatorliste og mapping
    flat_gen_ids = []
    gen_to_node_type = {}

    for node, gens, types in zip(Nodes, gen_idx, gen_type):
        for gen, typ in zip(gens, types):
            flat_gen_ids.append(gen)
            gen_to_node_type[gen] = (node, typ)

    print(f"🔍 Henter produksjon for {len(flat_gen_ids)} generatorer fra databasen...")

    # === 2. Hent produksjon for ALLE generatorer i én spørring ===
    generator_outputs = database.getResultGeneratorPowerPerGenerator(flat_gen_ids, (start_hour, end_hour))

    # === 3. Strukturér produksjon per node og type ===
    production_per_node = defaultdict(lambda: defaultdict(list))

    for gen_id, output in generator_outputs.items():
        node, typ = gen_to_node_type[gen_id]
        production_per_node[node][typ].append(output)

    print(f"✅ Ferdig strukturert produksjon for {len(production_per_node)} noder.")

    return dict(production_per_node), gen_idx, gen_type





# def GetProductionAtSpecificNodes(Nodes, data: GridData, database: Database, start_hour, end_hour):
#     """
#     Henter produksjonsdata for spesifikke noder i et gitt tidsintervall.
#
#     Args:
#         Nodes (list): Liste over nodenavn (f.eks. ['NO1_1', 'NO1_2']).
#         data (object): Datastruktur med informasjon om noder, generatorer osv.
#         database (object): Databaseforbindelse for å hente produksjonsdata.
#         start_hour (int): Startindeks for tidsserien.
#         end_hour (int): Sluttindeks for tidsserien.
#
#     Returns:
#         tuple:
#             - production_per_node (dict): Produksjonsdata per node og type.
#             - gen_idx (list): Liste over generator-IDer per node.
#             - gen_type (list): Liste over generatortyper per node.
#     """
#     print(f'Collecting Production at Nodes {", ".join(Nodes)}')
#
#     # === FINN INDEKSENE FOR NODENE ===
#     node_idx = [int(data.node[data.node['id'] == node].index[0]) for node in Nodes]
#
#     # === HENT GENERATORER OG DERES TYPER ===
#     gen_idx = [[gen for gen in data.getGeneratorsAtNode(idx)] for idx in node_idx]
#     gen_type = [[data.generator.loc[gen, "type"] for gen in gens] for gens in gen_idx]
#
#     flat_gen_idx = [gen for sublist in gen_idx for gen in sublist]  # Flater ut listen
#
#     # === HENT PRODUKSJONSDATA FRA DATABASE ===
#     power_output = {gen: database.getResultGeneratorPower([gen], (start_hour, end_hour)) for gen in flat_gen_idx}
#     # power_output = {gen: database.getResultGeneratorPower([gen], (start_hour, end_hour)) for gen in flat_gen_idx}
#
#     # === ORGANISERE PRODUKSJON PER NODE OG TYPE ===
#     production_per_node = {node: {} for node in Nodes}
#     for node, gen_list, type_list in zip(Nodes, gen_idx, gen_type):
#         for gen, typ in zip(gen_list, type_list):
#             production_per_node[node].setdefault(typ, []).append(power_output.get(gen, [1]))  # Setter 0 hvis data mangler
#
#     return production_per_node, gen_idx, gen_type



#### Calculate capture price

def CalculateCapturePrice(production_per_node, nodal_prices_per_node):
    result_capture_price = []
    result_capture_rate = []

    for node, types in production_per_node.items():
        prices = np.array(nodal_prices_per_node[node])
        avg_price = prices.mean()
        node_capture_price = {"Node": node}
        node_capture_rate = {"Node": node}

        for typ, all_production_series in types.items():
            # Vectoriser summen direkte med numpy
            total_production = np.sum(all_production_series, axis=0)

            total_production_sum = total_production.sum()
            if total_production_sum > 0:
                capture_price = np.sum(total_production * prices) / total_production_sum
                node_capture_price[f"{typ.capitalize()} (EUR/MWh)"] = capture_price
                node_capture_rate[f"{typ.capitalize()} (Capture Rate)"] = capture_price / avg_price
            else:
                node_capture_price[f"{typ.capitalize()} (EUR/MWh)"] = None
                node_capture_rate[f"{typ.capitalize()} (Capture Rate)"] = None

        result_capture_price.append(node_capture_price)
        result_capture_rate.append(node_capture_rate)

    return pd.DataFrame(result_capture_price).set_index("Node"), pd.DataFrame(result_capture_rate).set_index("Node")


def CalculateCapturePriceOverYears(START, END, nodes, data, database, timezone):
    all_years_capture_prices = []

    for year in range(START["year"], END["year"] + 1):
        period_start = {"year": year, "month": 1, "day": 1, "hour": 0}
        period_end = {"year": year, "month": 12, "day": 31, "hour": 23}

        start_hour, end_hour = get_hour_range(year, year, timezone, period_start, period_end)

        production_per_node, _, _ = GetProductionAtSpecificNodes(
            nodes, data, database, start_hour, end_hour
        )
        nodal_prices_per_node = GetPriceAtSpecificNodes(
            nodes, data, database, start_hour, end_hour
        )

        capture_price_df, _ = CalculateCapturePrice(production_per_node, nodal_prices_per_node)

        capture_price_df["Year"] = year
        capture_price_df["Node"] = capture_price_df.index
        all_years_capture_prices.append(capture_price_df)

    if not all_years_capture_prices:
        raise ValueError("No valid capture price data collected for any year.")

    full_df = pd.concat(all_years_capture_prices)
    full_df.set_index(["Year", "Node"], inplace=True)
    return full_df


def CalculateValueFactorOverYears(START, END, nodes, data, database, timezone):
    all_years_value_factors = []

    for year in range(START["year"], END["year"] + 1):
        period_start = {"year": year, "month": 1, "day": 1, "hour": 0}
        period_end = {"year": year, "month": 12, "day": 31, "hour": 23}

        start_hour, end_hour = get_hour_range(year, year, timezone, period_start, period_end)

        nodal_prices_per_node = GetPriceAtSpecificNodes(
            nodes, data, database, start_hour, end_hour
        )

        # Gjennomsnittspris per node for dette året
        avg_prices = {node: pd.Series(prices).mean() for node, prices in nodal_prices_per_node.items()}

        # Capture-priser for dette året
        production_per_node, _, _ = GetProductionAtSpecificNodes(
            nodes, data, database, start_hour, end_hour
        )
        capture_price_df, _ = CalculateCapturePrice(production_per_node, nodal_prices_per_node)

        # Beregn Value Factor per node
        vf_df = capture_price_df.copy()
        for node in vf_df.index:
            for col in vf_df.columns:
                vf_df.loc[node, col] = (
                    vf_df.loc[node, col] / avg_prices[node]
                    if pd.notnull(vf_df.loc[node, col]) else None
                )

        vf_df["Year"] = year
        vf_df["Node"] = vf_df.index
        all_years_value_factors.append(vf_df)

    full_vf_df = pd.concat(all_years_value_factors)
    full_vf_df.set_index(["Year", "Node"], inplace=True)
    return full_vf_df

def CalculateCapturePriceAndValueFactorOverYears(START, END, nodes, data, database, timezone):
    capture_price_dfs = []
    value_factor_dfs = []

    for year in range(START["year"], END["year"] + 1):
        period_start = {"year": year, "month": 1, "day": 1, "hour": 0}
        period_end = {"year": year, "month": 12, "day": 31, "hour": 23}
        start_hour, end_hour = get_hour_range(year, year, timezone, period_start, period_end)

        print(f"Fetching data for {year}")
        t0 = time.time()
        production_per_node, _, _ = GetProductionAtSpecificNodes(nodes, data, database, start_hour, end_hour)
        nodal_prices_per_node = GetPriceAtSpecificNodes(nodes, data, database, start_hour, end_hour)

        avg_prices = {node: pd.Series(prices).mean() for node, prices in nodal_prices_per_node.items()}
        capture_df, _ = CalculateCapturePrice(production_per_node, nodal_prices_per_node)

        # Append capture price
        df_cp = capture_df.copy()
        df_cp["Year"] = year
        df_cp["Node"] = df_cp.index
        capture_price_dfs.append(df_cp)

        # Create value factor
        df_vf = capture_df.copy()
        for node in df_vf.index:
            for col in df_vf.columns:
                df_vf.loc[node, col] = (
                    df_vf.loc[node, col] / avg_prices[node]
                    if pd.notnull(df_vf.loc[node, col]) else None
                )

        # Fjern (EUR/MWh) fra kolonnenavn
        df_vf.columns = [
            col.replace(" (EUR/MWh)", "") if "(EUR/MWh)" in col else col
            for col in df_vf.columns
        ]

        df_vf["Year"] = year
        df_vf["Node"] = df_vf.index
        value_factor_dfs.append(df_vf)

        t1 = time.time()
        print(f"Time taken for year {year}: {t1 - t0:.2f} seconds")

    # Kombiner alt
    capture_price_full = pd.concat(capture_price_dfs).set_index(["Year", "Node"])
    value_factor_full = pd.concat(value_factor_dfs).set_index(["Year", "Node"])

    return capture_price_full, value_factor_full













def GetConsumptionAtSpecificNodes(Nodes, data: GridData, database: Database, start_hour, end_hour):
    """
    Henter forbruksdata for spesifikke noder i et gitt tidsintervall.

    Args:
        Nodes (list): Liste over nodenavn (f.eks. ['NO1_1', 'NO1_2']).
        data (object): Datastruktur med informasjon om noder.
        database (object): Databaseforbindelse for å hente forbruksdata.
        start_hour (int): Startindeks for tidsserien.
        end_hour (int): Sluttindeks for tidsserien.

    Returns:
        dict: Forbruksdata per node med kategoriene "fixed", "flex" og "sum".
    """
    print(f'Collecting Consumption at Nodes {", ".join(Nodes)}')

    # === FINN INDEKSENE FOR NODENE ===
    node_idx = data.node[data.node['id'].isin(Nodes)].index.tolist()

    # === HENT FORBRUKSDATA FOR HVER NODE ===
    consumption_per_node = {node: {} for node in Nodes}

    for node, idx in zip(Nodes, node_idx):
        area = data.node.loc[idx, "area"]  # Finn område for noden
        demand_data = getDemandPerNodeFromDB(data, database, area, node, (start_hour, end_hour))

        consumption_per_node[node]["fixed"] = demand_data["fixed"]
        consumption_per_node[node]["flex"] = demand_data["flex"]
        consumption_per_node[node]["sum"] = demand_data["sum"]

    return consumption_per_node


def GetPriceAtSpecificNodes(Nodes, data: GridData, database: Database, start_hour, end_hour):
    print(f'💰 Henter priser for {len(Nodes)} noder i batch...')

    node_idx_map = {node: int(data.node[data.node['id'] == node].index[0]) for node in Nodes}
    node_indices = list(node_idx_map.values())

    # Hent alle priser i én spørring
    prices_by_index = database.getResultNodalPricesPerNode(node_indices, (start_hour, end_hour))

    # Map indeks tilbake til nodenavn
    inverse_index_map = {v: k for k, v in node_idx_map.items()}
    nodal_prices = {inverse_index_map[idx]: price_list for idx, price_list in prices_by_index.items()}

    return nodal_prices





# def GetPriceAtSpecificNodes(Nodes, data: GridData, database: Database, start_hour, end_hour):
#     """
#     Henter nodalpris for spesifikke noder i et gitt tidsintervall.
#
#     Args:
#         Nodes (list): Liste over nodenavn.
#         data (object): Datastruktur med informasjon om noder.
#         database (object): Databaseforbindelse for å hente priser.
#         start_hour (int): Startindeks for tidsserien.
#         end_hour (int): Sluttindeks for tidsserien.
#
#     Returns:
#         dict: Nodalpris per node.
#     """
#     print(f'Collecting Prices at Nodes {", ".join(Nodes)}')
#
#     # === FINN INDEKSENE FOR NODENE ===
#     node_idx = [int(data.node[data.node['id'] == node].index[0]) for node in Nodes]
#
#     # === HENT NODALPRIS FOR HVER NODE ===
#     nodal_prices = {node: database.getResultNodalPrice(idx, (start_hour, end_hour)) for node, idx in zip(Nodes, node_idx)}
#
#     return nodal_prices






def ExportToExcel(Nodes, production_per_node, consumption_per_node, nodal_prices_per_node, reservoir_filling_per_node, storage_cap, flow_data, START, END, case, version, OUTPUT_PATH):
    """
    Eksporterer produksjons-, forbruks-, fyllingsgrads- og nodalprisdata til en Excel-fil.

    Args:
        Nodes (list): Liste over nodenavn.
        production_per_node (dict): Produksjonsdata per node og type.
        consumption_per_node (dict): Forbruksdata per node.
        nodal_prices_per_node (dict): Nodalpriser per node.
        reservoir_filling_per_node (dict): Reservoarfylling per node.
        flow_data (DataFrame): Flow data for valgte linjer
        START (dict): Starttidspunkt som dictionary (f.eks. {"year": 2019, "month": 5, "day": 1, "hour": 12}).
        END (dict): Sluttidspunkt som dictionary (f.eks. {"year": 2019, "month": 6, "day": 1, "hour": 12}).
        case (str): Navn på caset.
        version (str): Versjonsnummer.
        OUTPUT_PATH: Filsti for lagring.

    Returns:
        str: Filnavn på den lagrede Excel-filen.
    """
    print(f'Collecting Data to Excel')

    # Konverter START og END til datetime-objekter
    start_datetime = datetime(START["year"], START["month"], START["day"], START["hour"])
    end_datetime = datetime(END["year"], END["month"], END["day"], END["hour"])

    all_types = ["nuclear", "hydro", "biomass", "ror", "wind_on", "wind_off", "solar", "fossil_other", "fossil_gas"]

    # === GENERER TIDSSTEG BASERT PÅ get_hour_range() ===
    # time_stamps = [start_datetime + timedelta(hours=i) for i in range(int((end_datetime - start_datetime).total_seconds() // 3600) + 1)]
    datetime_range = pd.date_range(start=start_datetime, end=end_datetime, freq='h', inclusive='left')


    # === GENERER FILNAVN ===
    # timestamp = datetime.now().strftime("%Y-%m-%d")
    start_str = start_datetime.strftime("%Y-%m-%d-%H")
    end_str = end_datetime.strftime("%Y-%m-%d-%H")
    filename = f"Prod_demand_nodes_{case}_{version}_{start_str}_to_{end_str}.xlsx"

    # === OPPRETT NY WORKBOOK ===
    wb = Workbook()

    for node in Nodes:
        # === OPPRETT ARKFANER FOR HVER NODE ===
        ws_production = wb.create_sheet(f"Production {node}")
        ws_consumption = wb.create_sheet(f"Consumption {node}")
        ws_price = wb.create_sheet(f"Price {node}")
        ws_reservoir = wb.create_sheet(f"Reservoir {node}")

        # === LEGG TIL OVERSKRIFTER ===
        ws_production.append(["Timestamp"] + all_types)
        ws_consumption.append(["Timestamp", "Fixed", "Flexible", "Consumption"])
        ws_price.append(["Timestamp", "Nodal Price"])
        ws_reservoir.append(["Timestamp", "Reservoir Filling", "Max storage capacity", "Reservoir Filling [%]"])


        # === FYLL PRODUKSJONSARKET ===
        for t, timestamp in enumerate(datetime_range):
            row = [timestamp.strftime("%Y-%m-%d %H:%M")]  # Formater tid riktig
            for typ in all_types:
                values = production_per_node[node].get(typ, [[0]])[0]  # Fjern dobbel liste-nesting
                value = values[t] if t < len(values) else 0  # Hent riktig indeks eller sett 0
                row.append(value)
            ws_production.append(row)
        auto_adjust_column_width(ws_production)

        # === FYLL FORBRUKSARKET ===
        for t, timestamp in enumerate(datetime_range):
            fixed = consumption_per_node[node]["fixed"]
            flex = consumption_per_node[node]["flex"]
            total = consumption_per_node[node]["sum"]

            ws_consumption.append([
                timestamp.strftime("%Y-%m-%d %H:%M"),
                fixed[t] if t < len(fixed) else 0,
                flex[t] if t < len(flex) else 0,
                total[t] if t < len(total) else 0,
            ])
        auto_adjust_column_width(ws_consumption)

        # === FYLL NODALPRISARKET ===
        for t, timestamp in enumerate(datetime_range):
            nodal_price = nodal_prices_per_node[node]  # Hent liste over nodalpriser for noden
            price_value = nodal_price[t] if t < len(nodal_price) else 0  # Hent pris eller sett 0 hvis ikke nok data

            ws_price.append([
                timestamp.strftime("%Y-%m-%d %H:%M"),
                price_value,
            ])
        auto_adjust_column_width(ws_price)

        # === FYLL RESERVOARFYLLINGSARKET ===
        for t, timestamp in enumerate(datetime_range):
            # Henter og pakker ut reservoardata
            if node in reservoir_filling_per_node:
                reservoir_data = reservoir_filling_per_node.get(node, [])
                if isinstance(reservoir_data, list):  # Sikrer riktig format
                    reservoir_values = [sum(x) for x in zip(*reservoir_data)]  # Slår sammen flere generatorer
                else:
                    reservoir_values = [0] * len(datetime_range)  # Hvis data er feil format
            else:
                reservoir_values = [0] * len(datetime_range)  # Hvis noden ikke finnes

            # Hent riktig verdi fra tidsserien eller sett 0
            reservoir_value = reservoir_values[t] if t < len(reservoir_values) else 0

            # Hent maks kapasitet for noden
            max_capacity = storage_cap.get(node, 0)  # Standard 0 hvis ikke funnet

            # Beregn fyllingsgrad (%)
            filling_percentage = (reservoir_value / max_capacity) * 100 if max_capacity > 0 else 0

            ws_reservoir.append([
                timestamp.strftime("%Y-%m-%d %H:%M"),
                round(reservoir_value, 4),  # Rund av til 4 desimaler
                round(max_capacity, 4) if max_capacity > 0 else "",  # Tom celle hvis ikke kapasitet
                round(filling_percentage, 8)  # Rund av til 8 desimaler
            ])
        auto_adjust_column_width(ws_reservoir)

    # === LEGG TIL FLOW-DATA HVIS TILGJENGELIG ===
    if flow_data is not None:
        for idx, row in flow_data.iterrows():
            from_node = row['from']
            to_node = row['to']
            sheet_name = f"Flow {from_node} → {to_node}"[:31]  # Excel begrensning

            # Forsøk å laste liste
            try:
                load_list = eval(row['load [MW]']) if isinstance(row['load [MW]'], str) else row['load [MW]']
            except Exception as e:
                print(f"❌ Hopper over {from_node} → {to_node}: feil i lesing av data ({e})")
                continue

            if len(load_list) != len(datetime_range):
                print(
                    f"⚠️ Hopper over {from_node} → {to_node}: Lengdemismatch ({len(load_list)} vs {len(datetime_range)})")
                continue

            if sheet_name in wb.sheetnames:
                print(f"⚠️ Ark {sheet_name} finnes allerede – hopper over.")
                continue

            # Lag ark og legg inn data
            ws_flow = wb.create_sheet(sheet_name)
            ws_flow.append(["Timestamp", "Load [MW]"])

            for t, timestamp in enumerate(datetime_range):
                ws_flow.append([
                    timestamp.strftime("%Y-%m-%d %H:%M"),
                    load_list[t]
                ])
            auto_adjust_column_width(ws_flow)

    # Fjern default ark
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # === LAGRE FIL ===
    filepath = pathlib.Path(OUTPUT_PATH) / filename
    wb.save(filepath)

    print(f"\nExcel-fil '{filename}' er lagret i {OUTPUT_PATH}!")

    return filename


def GetReservoirFillingAtSpecificNodes(Nodes, data: GridData, database: Database, start_hour, end_hour):
    """
    Henter reservoarfylling og maksimal kapasitet for spesifikke noder.
    """
    print(f'Collecting Reservoir Filling at Nodes {", ".join(Nodes)}')

    # === FINN INDEKSENE FOR NODENE ===
    node_idx = [int(data.node[data.node['id'] == node].index[0]) for node in Nodes]

    # === HENT LAGRINGSENHETER OG DERES KAPASITET ===
    storage_data = data.generator[
        (data.generator["node"].isin(Nodes)) &
        (data.generator["storage_cap"] > 0) &
        (data.generator["type"].isin(["hydro", "ror"]))
    ][["node", "storage_cap"]]

    storage_idx = storage_data.groupby("node").apply(lambda x: list(x.index)).to_dict()
    storage_cap = storage_data.groupby("node")["storage_cap"].sum().to_dict()  # Summerer kapasitet for hver node

    flat_storage_idx = [gen for sublist in storage_idx.values() for gen in sublist]

    # === HENT RESERVOARFYLLINGSNIVÅ FRA DATABASE ===
    storage_filling = {gen: database.getResultStorageFilling(gen, (start_hour, end_hour)) for gen in flat_storage_idx}

    # === ORGANISERE DATA ===
    reservoir_filling_per_node = {node: [] for node in Nodes}

    for node, gen_list in storage_idx.items():
        node_values = []
        for gen in gen_list:
            node_values.append(storage_filling.get(gen, [0]))  # Hent fyllingsdata eller 0
        reservoir_filling_per_node[node] = node_values

    return reservoir_filling_per_node, storage_cap



# === FLOW TO EXCEL ===


def getFlowDataOnBranches(data: GridData, db: Database, time_max_min, chosen_connections):
    """
    Collect flow on chosen connections.
    :param db:
    :param time_max_min:
    :param grid_data_path:
    :param chosen_connections:
    :return: flow_df
    """

    if chosen_connections is None:
        print('No connections chosen.')
        return pd.DataFrame()
    print(f'Collecting Flow Data at Lines {", ".join([f"{f} → {t}" for f, t in chosen_connections])}')

    if chosen_connections is not None:
        AC_interconnections, DC_interconnections = filter_connections_by_list(data, chosen_connections)
        AC_interconnections_capacity = AC_interconnections['capacity']
        DC_interconnections_capacity = DC_interconnections['capacity']

        # Get connections
        AC_dict, DC_dict = get_connections(data, chosen_connections)

        # Collect AC and DC flow data
        flow_data_AC = collect_flow_data(db, time_max_min, AC_dict, AC_interconnections_capacity, ac=True)
        flow_data_DC = collect_flow_data(db, time_max_min, DC_dict, DC_interconnections_capacity, ac=False)

        # Combine data into a single DataFrame
        flow_df = pd.concat([
            pd.DataFrame(flow_data_AC),
            pd.DataFrame(flow_data_DC)
        ], ignore_index=True)

    return flow_df


def writeFlowToExcel(flow_data, START, END, OUTPUT_PATH, case, version):
    """
    Writes flow data to Excel spreadsheet
    Args:
        flow_data (pandas.DataFrame): flow data from getFlowDataOnBranches
        START (pandas.DataFrame): start time of flow data
        END (pandas.DataFrame): end time of flow data
        TIMEZONE (pandas.DataFrame): timezone of flow data
        OUTPUT_PATH (pathlib.Path): output path

    Returns:
        str: Excel spreadsheet path
    """
    # Konverter START og END til datetime-objekter
    start_datetime = datetime(START["year"], START["month"], START["day"], START["hour"])
    end_datetime = datetime(END["year"], END["month"], END["day"], END["hour"])

    datetime_range = pd.date_range(start=start_datetime, end=end_datetime, freq='h', inclusive='left')

    # Create filename with timestamp
    filename = f"flow_data_{case}_{version}.xlsx"
    full_path = OUTPUT_PATH / filename

    # Track if any sheet was written
    sheet_written = False

    with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
        for idx, row in flow_data.iterrows():
            from_node = row['from']
            to_node = row['to']
            sheet_name = f"{from_node} → {to_node}"[:31]

            # Load list safely
            try:
                load_list = eval(row['load [MW]']) if isinstance(row['load [MW]'], str) else row['load [MW]']
            except Exception as e:
                print(f"❌ Failed to parse load for {from_node} → {to_node}: {e}")
                continue

            # Length check
            if len(load_list) != len(datetime_range):
                print(
                    f"⚠️ Skipping {from_node} → {to_node}: Length mismatch ({len(load_list)} vs {len(datetime_range)})")
                continue

            # Write sheet
            flow_df = pd.DataFrame({
                "Datetime": datetime_range,
                "Load [MW]": load_list
            })
            flow_df.to_excel(writer, sheet_name=sheet_name, index=False)
            sheet_written = True
            worksheet = writer.sheets[sheet_name]

            # Auto-adjust column widths
            for idx, col in enumerate(flow_df.columns, 1):
                max_length = max(flow_df[col].astype(str).map(len).max(), len(col)) + 2
                worksheet.column_dimensions[get_column_letter(idx)].width = max_length

    if not sheet_written:
        print("❗️No valid sheets were written. Excel file was not saved.")
        return None

    print(f"\n✅ Excel file saved at: {full_path}")
    return str(full_path)



def get_zone_production_summary(SELECTED_NODES, START, END, TIMEZONE, SIM_YEAR_START, SIM_YEAR_END, data, database):
    '''
    Retrieves production data for the selected nodes over the specified time period,
    aggregates the production by zone and by type, converts the results to TWh,
    and merges selected production types into broader categories.

    Parameters:
        SELECTED_NODES (list or str): List of node IDs to include or "ALL" to select all nodes.
        START (dict): Dictionary defining the start time with keys "year", "month", "day", "hour".
        END (dict): Dictionary defining the end time with keys "year", "month", "day", "hour".
        TIMEZONE (str): Timezone name.
        SIM_YEAR_START (datetime): Start of simulation year.
        SIM_YEAR_END (datetime): End of simulation year.
        data (object): Data object containing node information.
        database (object): Database connection or access object for production data.

    Returns:
        zone_summed_df (pd.DataFrame): Aggregated production per original production type, in TWh.
        zone_summed_merged_df (pd.DataFrame): Aggregated production per merged production type, with total, in TWh.
    '''

    Nodes = data.node["id"].dropna().unique().tolist() if SELECTED_NODES == "ALL" else SELECTED_NODES
    start_hour, end_hour = get_hour_range(SIM_YEAR_START, SIM_YEAR_END, TIMEZONE, START, END)
    production_per_node, gen_idx, gen_type = GetProductionAtSpecificNodes(Nodes, data, database, start_hour, end_hour)

    zone_sums = {}

    for node, prodtypes in production_per_node.items():
        zone = node.split("_")[0]
        if zone not in zone_sums:
            zone_sums[zone] = {}

        for prodtype, values_list in prodtypes.items():
            if not values_list or not values_list[0]:
                prod_sum = 0
            else:
                values = values_list[0]
                prod_sum = sum(values)

            if prodtype not in zone_sums[zone]:
                zone_sums[zone][prodtype] = prod_sum
            else:
                zone_sums[zone][prodtype] += prod_sum

    zone_summed_df = pd.DataFrame(zone_sums).T
    zone_summed_df = zone_summed_df / 1e6  # Convert from MWh to TWh

    merge_mapping = {
        "Hydro": ["hydro", "ror"],
        "Nuclear": ["nuclear"],
        "Solar": ["solar"],
        "Thermal": ["fossil_gas", "fossil_other", "biomass"],
        "Wind Onshore": ["wind_on"],
        "Wind Offshore": ["wind_off"]
    }

    zone_summed_merged = {}

    for new_type, old_types in merge_mapping.items():
        zone_summed_merged[new_type] = zone_summed_df[old_types].sum(axis=1, skipna=True)

    zone_summed_merged_df = pd.DataFrame(zone_summed_merged)

    zone_summed_merged_df["Production total"] = zone_summed_merged_df.sum(axis=1)

    desired_order = ["Production total", "Hydro", "Nuclear", "Solar", "Thermal", "Wind Onshore", "Wind Offshore"]
    zone_summed_merged_df = zone_summed_merged_df[desired_order]

    return zone_summed_df, zone_summed_merged_df



def getDemandInAllZonesFromDB(data, database, time_Demand, START, END, TIMEZONE, OUTPUT_PATH=None):
    zones = [zone for zone in data.node.zone.unique().tolist() if zone != 'SINK']
    demand={}
    for zone in zones:
        demand[zone] = getDemandPerZoneFromDB(data, database, area=zone[0:2], zone=zone, timeMaxMin=time_Demand)['sum']
    # date_range = pd.date_range(start=pd.to_datetime(START), end=pd.to_datetime(END), freq='h')
    year_start = datetime(START['year'], START['month'], START['day'], START['hour'], 0, tzinfo=TIMEZONE)
    year_end = datetime(END['year'], END['month'], END['day'], END['hour'], 0, tzinfo=TIMEZONE)
    df_demand = pd.DataFrame(demand, index=pd.date_range(start=year_start, end=year_end, freq='h'))
    # df_demand['time'] = pd.date_range(start=year_start, end=year_end, freq='h')
    if OUTPUT_PATH is not None:
        df_demand.to_csv(OUTPUT_PATH / f'demand_all_zones_{year_start.year}_{year_end.year}.csv')
    return df_demand


def get_production_summary_full_period(data, database, time_Prod, START, END, TIMEZONE, level='zone'):
    '''
    Retrieves production data for the specified level (zone or node) over the given time period,
    returns production by level and by type, converts to TWh for zones, and merges selected
    production types into broader categories. Excludes SINK nodes/zones and saves results to
    an Excel file with separate sheets for merged and non-merged DataFrames.

    Parameters:
        data (object): Data object containing node information.
        database (object): Database connection or access object for production data.
        time_Prod (tuple): Time range for production data.
        START (dict): Dictionary defining the start time with keys "year", "month", "day", "hour".
        END (dict): Dictionary defining the end time with keys "year", "month", "day", "hour".
        OUTPUT_PATH (Path): Path to save output Excel file.
        level (str): 'zone' or 'node' to specify the aggregation level.

    Returns:
        summed_df (pd.DataFrame): Production per original production type, in TWh (zone) or MWh (node).
        summed_merged_df (pd.DataFrame): Production per merged production type with total, in TWh (zone) or MWh (node).
    '''
    # Validate level
    if level not in ['zone', 'node']:
        raise ValueError("Level must be 'zone' or 'node'")

    # Get list of nodes and filter out SINK
    Nodes = [n for n in data.node["id"].dropna().unique().tolist() if 'SINK' not in n]

    # Get production data
    production_per_node, gen_idx, gen_type = GetProductionAtSpecificNodes(Nodes, data, database, time_Prod[0], time_Prod[1])

    # Create time index
    start_time = datetime(START['year'], START['month'], START['day'], START['hour'], 0)
    end_time = datetime(END['year'], END['month'], END['day'], END['hour'], 0)
    time_index = pd.date_range(start=start_time, end=end_time, freq='h')
    num_timesteps = len(time_index)

    # Initialize dictionary to store time-series data
    production_data = {}

    # Process production data
    for node, prodtypes in production_per_node.items():
        # Extract key based on level (zone or node)
        key = node.split("_")[0] if level == 'zone' else node
        if key not in production_data:
            production_data[key] = {}

        for prodtype, values_list in prodtypes.items():
            # Handle empty or null values
            if not values_list or not values_list[0]:
                values = [0] * num_timesteps
            else:
                values = values_list[0]
                if len(values) != num_timesteps:
                    raise ValueError(
                        f"Production data for node {node}, type {prodtype} has incorrect length: {len(values)} vs {num_timesteps}")

            # Store or sum time-series data
            if prodtype not in production_data[key]:
                production_data[key][prodtype] = values
            else:
                production_data[key][prodtype] = [sum(x) for x in zip(production_data[key][prodtype], values)]

    # Convert to DataFrame with multi-level columns
    columns = pd.MultiIndex.from_tuples(
        [(key, prodtype) for key in production_data for prodtype in production_data[key]],
        names=[level.capitalize(), 'Production Type']
    )
    summed_df = pd.DataFrame(
        data=[[production_data[key][prodtype][t] for key in production_data for prodtype in production_data[key]]
              for t in range(num_timesteps)],
        index=time_index,
        columns=columns
    )

    # # Convert to TWh for zones
    # if level == 'zone':
    #     summed_df = summed_df / 1e6

    # Drop columns with all zero values
    # summed_df = summed_df.loc[:, (summed_df != 0).any(axis=0)]

    # Merge production types
    merge_mapping = {
        "Hydro": ["hydro", "ror"],
        "Nuclear": ["nuclear"],
        "Solar": ["solar"],
        "Thermal": ["fossil_gas", "fossil_other", "biomass"],
        "Wind Onshore": ["wind_on"],
        "Wind Offshore": ["wind_off"]
    }

    # Initialize merged DataFrame
    merged_data = {}
    for key in production_data:
        for new_type, old_types in merge_mapping.items():
            valid_types = [t for t in old_types if (key, t) in summed_df.columns]
            if valid_types:
                merged_data[(key, new_type)] = summed_df[key][valid_types].sum(axis=1, skipna=True)
                if merged_data[(key, new_type)].eq(0).all():
                    del merged_data[(key, new_type)]

        # Add total production
        merged_data[(key, "Production Total")] = summed_df[key].sum(axis=1, skipna=True)

    # Create merged columns
    merged_columns = pd.MultiIndex.from_tuples(
        [col for col in merged_data.keys()],
        names=[level.capitalize(), 'Production Type']
    )
    summed_merged_df = pd.DataFrame(
        data={col: merged_data[col] for col in merged_columns},
        index=time_index
    )

    # Convert to TWh for zones in merged DataFrame
    # if level == 'zone':
    #     summed_merged_df = summed_merged_df / 1e6

    # Drop columns with all zero values in merged DataFrame
    # summed_merged_df = summed_merged_df.loc[:, (summed_merged_df != 0).any(axis=0)]

    return summed_df, summed_merged_df

def save_production_to_excel(data, database, time_period, START, END, TIMEZONE, OUTPUT_PATH, VERSION, relative=True):
    '''
    Generates production summaries for both zone and node levels, and saves them to
    separate sheets in a single Excel file.

    Parameters:
        data (object): Data object containing node information.
        database (object): Database connection or access object for production data.
        time_Prod (tuple): Time range for production data.
        START (dict): Dictionary defining the start time with keys "year", "month", "day", "hour".
        END (dict): Dictionary defining the end time with keys "year", "month", "day", "hour".
        OUTPUT_PATH (Path): Path to save output Excel file.
    '''
    # Generate DataFrames for zone and node levels

    start_time = datetime(START['year'], START['month'], START['day'], START['hour'], 0)
    end_time = datetime(END['year'], END['month'], END['day'], END['hour'], 0)


    zone_summed_df, zone_summed_merged_df = get_production_summary_full_period(
        data, database, time_period, START, END, TIMEZONE, level='zone'
    )
    node_summed_df, node_summed_merged_df = get_production_summary_full_period(
        data, database, time_period, START, END, TIMEZONE, level='node'
    )

    zone_demand_df = getDemandInAllZonesFromDB(data, database, time_period, START, END, TIMEZONE, OUTPUT_PATH=None)

    node_demand = collectDemandForAllNodesAllTimeStepsFromDB(data, database, time_period)
    node_demand = pd.DataFrame(node_demand, index=pd.date_range(start=start_time, end=end_time, freq='h'))

    # Get zonal prices for the selected zones
    zonal_prices = getZonalPrices(data, database, time_period, START, END)

    Nodes = data.node[data.node.zone != 'SINK'].id
    nodal_prices_per_node = GetPriceAtSpecificNodes(Nodes, data, database, time_period[0], time_period[-1])
    nodal_prices_per_node = pd.DataFrame(nodal_prices_per_node, index=pd.date_range(start=start_time, end=end_time, freq='h'))

    df_zone_flows = getFlowBetweenAllZones(data, database, time_period, START, END)

    df_storagefilling = getStorageFilling(data, database, time_period, START, END, relative)

    # Define Excel file path

    excel_path = OUTPUT_PATH / f'production_summary_{VERSION}_{start_time.year}_{end_time.year}.xlsx'

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Dictionary of DataFrames and sheet names
    dfs = {
        'Zone_Production': zone_summed_df,
        'Zone_Production_Merged': zone_summed_merged_df,
        'Node_Production': node_summed_df,
        'Node_Production_Merged': node_summed_merged_df,
        'Zone_Demand': zone_demand_df,
        'Node_Demand': node_demand,
        'Zonal_Prices': zonal_prices,
        'Nodal_Prices': nodal_prices_per_node,
        'Zone_Flows': df_zone_flows,
        'Storage_Filling': df_storagefilling
    }

    # Write each DataFrame to a separate sheet
    for sheet_name, df in dfs.items():
        print(f"Writing sheet: {sheet_name}")
        ws = wb.create_sheet(title=sheet_name)

        if 'Production' in sheet_name:
            # Write single-row combined headers
            ws.cell(row=1, column=1).value = 'Time'
            if not df.columns.empty and isinstance(df.columns, pd.MultiIndex):
                # Combine Zone and Production Type into a single header
                for col_idx, (zone, prod_type) in enumerate(df.columns, start=2):
                    combined_header = f"{zone} {prod_type.replace(' ', '_')}"
                    ws.cell(row=1, column=col_idx).value = combined_header
            else:
                # Fallback for non-MultiIndex columns
                for col_idx, col_name in enumerate(df.columns, start=2):
                    ws.cell(row=1, column=col_idx).value = str(col_name)

        elif 'Demand' in sheet_name:
            # Write single-row combined headers
            ws.cell(row=1, column=1).value = 'Time'
            for col_idx, col_name in enumerate(df.columns, start=2):
                combined_header = f"{col_name} Load"
                ws.cell(row=1, column=col_idx).value = combined_header

        elif 'Prices' in sheet_name:
            # Write single-row combined headers
            ws.cell(row=1, column=1).value = 'Time'
            for col_idx, col_name in enumerate(df.columns, start=2):
                combined_header = f"{col_name} Price"
                ws.cell(row=1, column=col_idx).value = combined_header

        elif 'Flows' in sheet_name:
            # Write single-row combined headers
            ws.cell(row=1, column=1).value = 'Time'
            for col_idx, col_name in enumerate(df.columns, start=2):
                combined_header = f"{col_name} Flow"
                ws.cell(row=1, column=col_idx).value = combined_header

        elif 'Storage' in sheet_name:
            # Write single-row combined headers
            ws.cell(row=1, column=1).value = 'Time'
            for col_idx, col_name in enumerate(df.columns, start=2):
                combined_header = f"{col_name} Filling"
                ws.cell(row=1, column=col_idx).value = combined_header

        # Write index (time) as YYYY-MM-DD HH:MM strings and data
        for row_idx, (time, row_data) in enumerate(df.iterrows(), start=2):
            # Convert time to timezone-naive datetime
            if time.tzinfo is not None:
                time = time.replace(tzinfo=None)
            # Write time as a datetime object
            ws.cell(row=row_idx, column=1).value = time
            # Set cell format to display as YYYY-MM-DD HH:MM
            ws.cell(row=row_idx, column=1).number_format = 'yyyy-mm-dd hh:mm'
            for col_idx, value in enumerate(row_data, start=2):
                ws.cell(row=row_idx, column=col_idx).value = float(value) if pd.notnull(value) else 0

        print(f"Sheet {sheet_name} written with {df.shape[0]} rows and {df.shape[1]} columns")

    # Save workbook
    try:
        wb.save(excel_path)
        print(f"Excel file saved: {excel_path}")
    except Exception as e:
        print(f"Failed to save Excel file: {e}")
        raise


    # zone_summed_df.to_csv(OUTPUT_PATH / 'zone_summed_df.csv')
    # zone_summed_merged_df.to_csv(OUTPUT_PATH / 'zone_summed_merged_df.csv')
    # node_summed_df.to_csv(OUTPUT_PATH / 'node_summed_df.csv')
    # node_summed_merged_df.to_csv(OUTPUT_PATH / 'node_summed_merged_df.csv')


def getEnergyBalance(data, database, SIM_YEAR_START, SIM_YEAR_END, TIMEZONE, OUTPUT_PATH, VERSION, YEARS):
    energyBalance = {}

    for i in range(0, YEARS):
        year = 1991 + i
        print(year)

        START = {"year": year, "month": 1, "day": 1, "hour": 0}
        END = {"year": year, "month": 12, "day": 31, "hour": 23}
        time_EB = get_hour_range(SIM_YEAR_START, SIM_YEAR_END, TIMEZONE, START, END)
        all_nodes = data.node.id

        totalProduction = collectProductionForAllNodesFromDB(data, database, time_EB)
        flow_data = collectFlowDataOnALLBranches(data, database, time_EB)

        totalDemand = collectDemandForAllNodesFromDB(data, database, time_EB)
        # Remove all keys containing 'SINK'
        totalDemand = {key: value for key, value in totalDemand.items() if 'SINK' not in key}
        totalLoadShedding = database.getResultLoadheddingSum(timeMaxMin=time_EB)

        # Calculate energy balance at node and zone levels
        node_energyBalance = getEnergyBalanceNodeLevel(all_nodes, totalDemand, totalProduction, totalLoadShedding,
                                                       flow_data, OUTPUT_PATH / 'data_files/energy_balance', VERSION,
                                                       START)
        zone_energyBalance = getEnergyBalanceZoneLevel(all_nodes, totalDemand, totalProduction, totalLoadShedding,
                                                       flow_data, OUTPUT_PATH / 'data_files/energy_balance', VERSION,
                                                       START)
        # Store energy balance results in the dictionary
        energyBalance[year] = {
            "node_level": node_energyBalance,
            "zone_level": zone_energyBalance
        }

    # get average balance in each country
    # Initialize dictionary to store yearly sums for each country and metric
    sumBalance = {
        'NO': {'Production': [], 'Demand': [], 'Balance_mls': [], 'Balance_uls': []},
        'SE': {'Production': [], 'Demand': [], 'Balance_mls': [], 'Balance_uls': []},
        'FI': {'Production': [], 'Demand': [], 'Balance_mls': [], 'Balance_uls': []},
        'DK': {'Production': [], 'Demand': [], 'Balance_mls': [], 'Balance_uls': []},
        'GB': {'Production': [], 'Demand': [], 'Balance_mls': [], 'Balance_uls': []},
        'DE': {'Production': [], 'Demand': [], 'Balance_mls': [], 'Balance_uls': []},
        'NL': {'Production': [], 'Demand': [], 'Balance_mls': [], 'Balance_uls': []},
        'LT': {'Production': [], 'Demand': [], 'Balance_mls': [], 'Balance_uls': []},
        'PL': {'Production': [], 'Demand': [], 'Balance_mls': [], 'Balance_uls': []},
        'EE': {'Production': [], 'Demand': [], 'Balance_mls': [], 'Balance_uls': []},
    }
    countries = data.node.area.unique()
    countries = countries[countries != 'SINK'].tolist()

    all_zones = data.node.zone.unique()
    all_zones = all_zones[all_zones != 'SINK'].tolist()
    metrics = {
        'Balance_mls': 'Balance_mls',
        'Balance_uls': 'Balance_uls',
        'Production': 'Production',
        'Demand': 'Demand'
    }

    # Iterate over years and compute sums
    for year in energyBalance:
        df = energyBalance[year]['zone_level']  # DataFrame for the year
        for country in countries:
            # Filter zones for the country
            mask = df['Zone'].str.contains(country, na=False)
            for sum_key, df_col in metrics.items():
                # Sum the relevant column for filtered zones
                sumBalance[f'{country}'][sum_key].append(
                    df[mask][df_col].astype(float).fillna(0).sum()
                )

    # Calculate mean for each country and metric
    mean_balance = {
        f'{country}': {
            metric: sum(values) / len(values) if values else 0
            for metric, values in metrics_dict.items()
        }
        for country, metrics_dict in sumBalance.items()
    }
    mean_balance = pd.DataFrame(mean_balance).T

    return energyBalance, mean_balance



def getZonalPrices(data: GridData, database: Database, time_period, START, END):
    """
    Calculate zonal prices as the average of nodal prices within each zone, excluding SINK nodes.

    Args:
        data (GridData): Grid data containing node and zone information.
        database (Database): Database connection to fetch nodal prices.
        start_hour (int): Start index for the time range.
        end_hour (int): End index for the time range.

    Returns:
        dict: Zonal prices with zone names as keys and lists of average prices as values.
              Format: {zone: [avg_price_t1, avg_price_t2, ...]} for non-SINK zones.
    """
    print("💰 Collecting nodal prices for non-SINK nodes...")

    # Filter out SINK nodes
    non_sink_nodes = data.node[data.node['zone'] != 'SINK']

    # Map nodes to their respective zones
    zone_to_nodes = non_sink_nodes.groupby('zone')['id'].apply(list).to_dict()

    # Collect node indices for non-SINK nodes
    node_idx_map = {node: int(non_sink_nodes[non_sink_nodes['id'] == node].index[0]) for node in non_sink_nodes['id']}
    node_indices = list(node_idx_map.values())

    # Fetch all nodal prices in one query
    prices_by_index = database.getResultNodalPricesPerNode(node_indices, time_period)

    # Map prices back to nodes
    inverse_index_map = {v: k for k, v in node_idx_map.items()}
    nodal_prices = {inverse_index_map[idx]: prices for idx, prices in prices_by_index.items()}

    # Calculate zonal prices as time-series averages
    zonal_prices = {}
    for zone, nodes in zone_to_nodes.items():
        # Get prices for all nodes in the zone
        valid_prices = [nodal_prices[node] for node in nodes if node in nodal_prices]
        if not valid_prices:
            zonal_prices[zone] = [0] * (time_period[-1] - time_period[1] + 1)  # Fallback for empty zones
            continue

        # Convert to numpy array for efficient averaging
        prices_array = np.array(valid_prices)  # Shape: (n_nodes, n_timesteps)
        # Compute mean across nodes for each time step
        avg_prices = np.mean(prices_array, axis=0).tolist()  # Shape: (n_timesteps,)
        zonal_prices[zone] = avg_prices

    year_start = datetime(START['year'], START['month'], START['day'], START['hour'], 0)
    year_end = datetime(END['year'], END['month'], END['day'], END['hour'], 0)
    zonal_prices = pd.DataFrame(zonal_prices, index=pd.date_range(start=year_start, end=year_end, freq='h'))

    return zonal_prices



def process_windoff_sensitivity(data, database, time_period) -> pd.DataFrame:
    """
    Process sensitivity data for offshore wind generators, including nodal prices and derived metrics.

    Parameters:
    -----------
    data : object
        Data object with 'generator' (DataFrame with 'type', 'node' columns) and 'node' (DataFrame with 'id' column).
    database : object
        Database object with methods 'getResultGeneratorSens' and 'getResultNodalPricesMean'.
    sim_year_start : int
        Start year of the simulation.
    sim_year_end : int
        End year of the simulation.
    timezone : str
        Timezone for the time period calculation.
    start : dict
        Start time with keys 'year', 'month', 'day', 'hour'.
    end : dict
        End time with keys 'year', 'month', 'day', 'hour'.

    Returns:
    --------
    pd.DataFrame
        Sorted DataFrame with columns: 'generator_idx', 'node', 'sensitivity_eur',
        'sensitivity_avg_eur_h', 'nodal_price_avg_eur_mwh', 'sensitivity_diff_eur_mwh'.
        Returns empty DataFrame if no data is available or on error.
    """
    try:
        # Get offshore wind generator indices
        generator_idx = data.generator[data.generator.type == 'wind_off'].index.tolist()
        if not generator_idx:
            logging.warning("No offshore wind generators found.")
            return pd.DataFrame()

        # Get time period
        n_timesteps = time_period[1] - time_period[0]
        if n_timesteps <= 0:
            logging.error("Invalid time period: end time must be after start time.")
            return pd.DataFrame()

        # Retrieve and process sensitivity data
        df_sens = database.getResultGeneratorSens(time_period, generator_idx)
        if df_sens.empty:
            logging.warning("No sensitivity data retrieved.")
            return pd.DataFrame()

        # Sum sensitivities and create DataFrame
        df = df_sens.sum().to_frame(name='sensitivity_eur').reset_index().rename(columns={'indx': 'generator_idx'})
        df['node'] = df['generator_idx'].map(data.generator['node'])

        # Calculate average sensitivity per hour
        df['sensitivity_avg_eur_h'] = df['sensitivity_eur'] / n_timesteps

        # Map nodal prices
        node_ids = list(data.node['id'])
        avg_prices = database.getResultNodalPricesMean(time_period)
        if len(avg_prices) != len(node_ids):
            logging.warning(
                f"Number of prices ({len(avg_prices)}) does not match number of nodes ({len(node_ids)}). "
                f"Using first {min(len(node_ids), len(avg_prices))} nodes."
            )
        price_map = dict(zip(node_ids[:min(len(node_ids), len(avg_prices))], avg_prices[:min(len(node_ids), len(avg_prices))]))
        df['nodal_price_avg_eur_mwh'] = df['node'].map(price_map)

        # Calculate sensitivity difference
        df['sensitivity_diff_eur_mwh'] = df['nodal_price_avg_eur_mwh'] + df['sensitivity_avg_eur_h']

        # Select columns and sort by sensitivity
        df = df[['generator_idx', 'node', 'sensitivity_eur', 'sensitivity_avg_eur_h',
                 'nodal_price_avg_eur_mwh', 'sensitivity_diff_eur_mwh']]
        df = df.sort_values('sensitivity_eur', ascending=True)

        logging.info("Processed sensitivity data for %d generators.", len(generator_idx))
        return df

    except Exception as e:
        logging.error("Error processing sensitivity data: %s", e)
        return pd.DataFrame()


def generatorSensitivityRanking(data, database, generator_type, time_period, weighted=True):
    """
    Ranks generators of a specified type based on their sensitivity and inflow profile.

    Parameters:
    - generator_type (str): Type of generator to investigate (e.g., 'hydro', 'wind', 'solar')
    - data: Data object containing generator and profile information
    - database: Database object to retrieve sensitivity data
    - SIM_YEAR_START (int): Simulation start year
    - SIM_YEAR_END (int): Simulation end year
    - TIMEZONE (str): Timezone for the simulation
    - START (dict): Start time dictionary with year, month, day, hour
    - END (dict): End time dictionary with year, month, day, hour

    Returns:
    - pd.DataFrame: DataFrame with generator_idx, rank, and node, sorted by rank
    """
    # Get time period
    min_time, max_time = time_period  # Unpack min, max from time_period

    # Filter generators by type
    generator_idx = data.generator[data.generator.type == generator_type].index.tolist()

    # Get inflow ref to generator
    generator_inflow = data.generator.inflow_ref[generator_idx].tolist()
    gen_to_inflow_map = dict(zip(generator_idx, generator_inflow))

    # Get inflow profile to generator
    generator_inflow_profile = data.profiles[data.generator.inflow_ref[generator_idx].unique().tolist()].loc[
                               min_time:max_time]

    # Retrieve and process sensitivity data
    df_sens = database.getResultGeneratorSens(time_period, generator_idx)

    # Retrieve average nodal prices
    avg_prices = database.getResultNodalPricesMean(time_period)
    node_ids = data.generator.node.unique()  # Assuming node_ids come from unique nodes in generator data
    price_map = dict(zip(node_ids[:min(len(node_ids), len(avg_prices))], avg_prices[:min(len(node_ids), len(avg_prices))]))

    # Initialize results
    ranks = []

    for gen_idx, inflow_ref in gen_to_inflow_map.items():
        sens = df_sens[gen_idx]
        inflow = generator_inflow_profile[inflow_ref]

        numerator = (inflow * sens).sum() if weighted else sens.sum()
        denominator = inflow.sum() if weighted else len(sens)


        # Avoid division by zero
        if denominator == 0:
            rank = np.nan
            print(f"Warning: Denominator is zero for generator {gen_idx}. Rank set to NaN.")
        else:
            rank = numerator / denominator

        ranks.append({'generator_idx': gen_idx, 'sens': rank})

    # Create and format output DataFrame
    df = pd.DataFrame(ranks)
    df['node'] = df['generator_idx'].map(data.generator['node'])
    df['avg_nodal_price'] = df['node'].map(price_map)
    df = df[['generator_idx', 'node', 'sens', 'avg_nodal_price']]
    df = df.sort_values('sens', ascending=True)

    return df



def generatorSensitivityRankingALL(data, database, time_period, OUTPUT_PATH_PLOTS, inflow_weighted=True, save_fig=False,
                                   include_fliers=False, area_filter=None):
    """
    Computes the normalized sensitivity of generators and ranks them by type.
    This function retrieves the dual sensitivities for all generators over a specified time period,
    calculates the normalized sensitivity, and generates a boxplot of the results.

    Parameters:
    - data: The grid data object containing generator and profile information.
    - database: The database object to retrieve results from.
    - gen_type: The type of generator to filter by (e.g., 'wind_off', 'solar').
    - time_period: The time period for which to compute sensitivities.
    - inflow_weighted: If True, the sensitivity is weighted by the inflow profile.
    - save_fig: If True, saves the generated plot to a file.
    - include_fliers: If True, includes outliers in the boxplot.
    - area_filter: Optional string to filter generators by area (e.g., "NO1"). If None, all generators are included.

    """
    min_time, max_time = time_period
    # Get all generator indices and types
    all_gens = data.generator.index.tolist()

    if area_filter is not None:
        # Filter generator indices based on area (e.g., "NO1")
        all_gens = data.generator[data.generator.node.str.startswith(area_filter)].index.tolist()

    if len(all_gens) == 0:
        print(f"No generators found in area '{area_filter}'.")
        return pd.DataFrame()

    gen_types = data.generator['type']

    # Retrieve dual sensitivities for all generators over the time period
    df_sens = database.getResultGeneratorSens(time_period, all_gens)

    # Get inflow profile mapping
    generator_inflow_refs = data.generator.loc[all_gens, 'inflow_ref']
    gen_to_inflow_map = dict(zip(all_gens, generator_inflow_refs))

    # Load inflow profiles
    unique_inflows = generator_inflow_refs.unique().tolist()
    inflow_profiles = data.profiles[unique_inflows].loc[min_time:max_time]

    # Compute normalized sensitivity (R_i) for each generator
    ranks = []
    for gen_idx, inflow_ref in gen_to_inflow_map.items():
        sens = df_sens[gen_idx].abs()
        inflow = inflow_profiles[inflow_ref]

        numerator = (inflow * sens).sum()
        denominator = inflow.sum()

        if denominator == 0:
            rank = np.nan
            print(f"Warning: Denominator is zero for generator {gen_idx}. Rank set to NaN.")
        else:
            if inflow_weighted:
                # Normalized sensitivity
                rank = numerator / denominator
            else:
                # Non-weighted sensitivity
                rank = sens.mean()

        ranks.append({
            'generator_idx': gen_idx,
            'sens': rank,
            'type': gen_types.loc[gen_idx],
            'node': data.generator.loc[gen_idx, 'node']
        })

    # Create DataFrame of results
    df_sens_ranked = pd.DataFrame(ranks).dropna(subset=['sens'])

    # Now you can group by type and create a boxplot
    import seaborn as sns
    import matplotlib.pyplot as plt

    # Define a color palette that matches the nature of each generator type
    custom_palette = {
        'wind_off': '#1f77b4',    # Blue-ish for offshore wind
        'wind_on': '#1f77b4',     # Same tone for onshore wind
        'solar': '#ff7f0e',       # Orange for solar
        'hydro': '#2ca02c',       # Green for hydro
        'ror': '#17becf',         # Light blue for run-of-river
        'biomass': '#8c564b',     # Brown/earthy for biomass
        'fossil_gas': '#7f7f7f',  # Gray for fossil
        'fossil_other': '#6E5849',  # Dark brown for other fossil fuels
        'nuclear': '#9467bd',     # Purple for nuclear
    }


    # Step 1: Calculate the median sensitivity per generator type
    type_order = (
        df_sens_ranked.groupby('type')['sens']
        .median()
        .abs()                   # In case you've taken absolute values
        .sort_values(ascending=False)
        .index.tolist()
    )

    plt.figure(figsize=(12, 6))

    sns.boxplot(
        data=df_sens_ranked,
        x='type',
        y='sens',
        palette=custom_palette,
        order=type_order,        # Apply the custom order
        showfliers=include_fliers
    )
    title_area = f" ({area_filter})" if area_filter else ""
    title_weight = "Inflow Weighted" if inflow_weighted else "Not Weighted"
    plt.title(f'Generator Sensitivities by Type{title_area} ({title_weight})')
    plt.ylabel('Normalized Sensitivity EUR/MW')
    plt.xlabel('Generator Type')
    plt.xticks(rotation=45)
    plt.tight_layout()
    if save_fig:
        area_tag = f"_{area_filter}" if area_filter else ""
        weight_tag = "inflow_weighted" if inflow_weighted else "not_weighted"
        filename = f'gen_sens_by_type{area_tag}_{weight_tag}_{VERSION}.pdf'
        plt.savefig(OUTPUT_PATH_PLOTS / filename)
    plt.show()
    return df_sens_ranked


def getFlowBetweenAllZones(data, database, time_period, START, END):
    print(f'🔄 Collecting flow data between all zones from {START['year']} to {END['year']}...')

    start_time = datetime(START['year'], START['month'], START['day'], START['hour'], 0)
    end_time = datetime(END['year'], END['month'], END['day'], END['hour'], 0)
    time_index = pd.date_range(start=start_time, end=end_time, freq='h')
    flow_data = collectFlowDataOnALLBranches(data, database, time_period)

    all_nodes = data.node.id.tolist()  # Get all node names from the grid data
    node_names = [node for node in all_nodes]

    # Create node-to-zone mapping
    node_to_zone = {}
    for node in node_names:
        if '_' in node:
            # Take prefix before first underscore (e.g., 'DK1_3' -> 'DK1', 'SE3_hub_east' -> 'SE3')
            zone = node.split('_')[0]
        else:
            # No underscore (e.g., 'GB', 'DE') -> use full name as zone
            zone = node
        node_to_zone[node] = zone

    # Assuming flow_data is a list of lists or a DataFrame
    # Convert to DataFrame for easier handling if not already
    if isinstance(flow_data, list):
        flow_df = pd.DataFrame(flow_data[1:], columns=flow_data[0])
    else:
        flow_df = flow_data

    # Dictionary to hold time series flow between zones
    zone_to_zone_flow = defaultdict(lambda: np.zeros(len(time_period)))
    line_flows = defaultdict(list)

    # Optional: Set to hold all unique zone connections
    zone_connections = set()
    import ast
    # Process each line in flow_data
    for _, row in flow_df.iterrows():
        from_node = row['from']
        to_node = row['to']
        loads = row['load [MW]']  # List of load values

        # Convert load to array
        if isinstance(loads, str):
            loads = ast.literal_eval(loads)
        loads = np.array(loads).flatten()

        # if loads.shape[0] != 8760:
        #     print(f"Invalid load shape for line {from_node}-{to_node}: {loads.shape}")
        #     continue

        line_flows[(from_node, to_node)].append(loads)

        from_zone = node_to_zone.get(from_node)
        to_zone = node_to_zone.get(to_node)

        # Skip if same zone or undefined
        if from_zone is None or to_zone is None or from_zone == to_zone:
            continue

        # Always sort zone names to make the key symmetric
        key = tuple(sorted([from_zone, to_zone]))
        zone_connections.add(key)

        # Determine sign: +1 if direction matches key order, -1 otherwise
        direction = 1 if (from_zone, to_zone) == key else -1

        # Add signed flow
        existing = zone_to_zone_flow[key]
        if not isinstance(existing, np.ndarray) or existing.shape[0] != 8760:
            existing = np.zeros(8760)

        zone_to_zone_flow[key] = existing + direction * loads

    # Convert keys to single string: "SE2→SE3"
    flat_columns = {f"{k[0]}-{k[1]}": v for k, v in zone_to_zone_flow.items()}

    df_zone_flows = pd.DataFrame.from_dict(flat_columns)
    # Set index
    df_zone_flows.index = time_index
    df_zone_flows.index.name = "Time"

    return df_zone_flows


def getStorageFilling(data, database, time_period, START, END, relative=True):
    """ Collects storage filling data for specified areas and zones. """
    areas = ['NO', 'SE', 'FI']
    storfilling = pd.DataFrame()
    print(f"💧 Collecting storage filling data for areas: {', '.join(areas)}")

    for area in areas:
        storfilling[area] = getStorageFillingInAreaFromDB(data=data,
                                                          db=database,
                                                          areas=[area],
                                                          generator_type=['hydro', 'ror'],
                                                          relative_storage=relative,
                                                          timeMaxMin=time_period)
        if relative:
            storfilling[area] = storfilling[area] * 100

    zones = ['NO1', 'NO2', 'NO3', 'NO4', 'NO5', 'SE1', 'SE2', 'SE3', 'SE4']
    print(f"💧 Collecting storage filling data for zones: {', '.join(zones)}")

    for zone in zones:
        storfilling[zone] = getStorageFillingInZoneFromDB(data=data,
                                                          db=database,
                                                          zones=[zone],
                                                          generator_type=['hydro', 'ror'],
                                                          relative_storage=relative,
                                                          timeMaxMin=time_period)
        if relative:
            storfilling[zone] = storfilling[zone] * 100
    # Compute the correct DATE_START for this year
    start_time = datetime(START['year'], START['month'], START['day'], START['hour'], 0)
    end_time = datetime(END['year'], END['month'], END['day'], END['hour'], 0)
    storfilling.index = pd.date_range(start=start_time, end=end_time, freq='h')
    storfilling.index.name = "Time"
    return storfilling
