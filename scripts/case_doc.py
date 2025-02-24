import pathlib
import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def read_data(case):
    base_path = pathlib.Path(f"../case_{case}/data/system")
    generator_file = base_path / "generator.csv"
    consumer_file = base_path / "consumer.csv"  # Legger til consumer-filen
    return pd.read_csv(generator_file), pd.read_csv(consumer_file)

def ensure_all_types_present(data, group_col, all_types):
    """
    Sikrer at alle typer er representert i hver gruppe.
    Hvis en type mangler, legges den til med pmax=0 og N/A.
    """
    groups = data[group_col].unique()
    complete_data = []

    for group in groups:
        group_data = data[data[group_col] == group]
        complete_data.append(group_data)  # Legg til eksisterende data
        missing_types = set(all_types) - set(group_data['type'])
        for missing_type in missing_types:
            # Legg til manglende type som en DataFrame
            missing_row = pd.DataFrame([{
                group_col: group,
                'type': missing_type,
                'pmax': 0,
                'fuelcost': pd.NA,
                'storage_cap': pd.NA,
                'storage_price': pd.NA,
                'inflow_fac': pd.NA
            }])
            complete_data.append(missing_row)  # Legg til manglende data

    # Kombinerer eksisterende data med de manglende typene
    return pd.concat(complete_data, ignore_index=True)

def process_grouped_data(data, group_col):
    """
    Behandler data for en spesifikk gruppering
    """
    # Sikrer at alle typer er representert i hver gruppe
    all_types = data['type'].unique()
    data = ensure_all_types_present(data, group_col, all_types)

    grouped = data.groupby([group_col, 'type']).agg({
        'pmax': 'sum',
        'fuelcost': 'mean',
        'storage_cap': lambda x: (x / 1e6).sum(),  # Konverter storage_cap til TWh
        'storage_price': 'mean',
        'inflow_fac': 'mean'
    }).reset_index()

    # Navneendring for metrics
    metric_names = {
        'pmax': 'Installed capacity [MW]',
        'fuelcost': 'Fuelcost [Eur/MWh]',
        'storage_cap': 'Storage capacity [TWh]',
        'storage_price': 'Storage price [Eur/MWh]',
        'inflow_fac': 'Inflow factor'
    }

    metrics = ['pmax', 'fuelcost', 'storage_cap', 'storage_price', 'inflow_fac']
    tables = []
    for metric in metrics:
        table = grouped.pivot(index='type', columns=group_col, values=metric).reset_index()
        table['Metric'] = metric_names[metric]  # Bruk det nye navnet for metrics
        table = table.fillna("N/A")  # Sett N/A der verdier mangler

        # Sett spesifikke metrics til N/A basert på betingelser
        if metric in ['fuelcost', 'storage_cap', 'storage_price', 'inflow_fac']:
            pmax_table = tables[0] if tables else grouped.pivot(index='type', columns=group_col,
                                                                values='pmax').reset_index()
            for col in table.columns[1:-1]:  # Skipper 'Type' og 'Metric'-kolonner
                if metric in ['fuelcost', 'inflow_fac']:
                    table.loc[pmax_table[col] == 0, col] = "N/A"  # Fuelcost og inflow_fac er irrelevant hvis pmax er 0
                elif metric == 'storage_price':
                    storage_cap_table = tables[2] if len(tables) > 2 else grouped.pivot(index='type', columns=group_col,
                                                                                        values='storage_cap').reset_index()
                    table.loc[
                        storage_cap_table[col] == 0, col] = "N/A"  # Storage price er irrelevant hvis storage_cap er 0

        tables.append(table)

    # Kombiner og legg til Type-kolonnen
    combined = pd.concat(tables, ignore_index=True)
    combined.rename(columns={'type': 'Type'}, inplace=True)

    # Rekkefølge for kolonner
    cols = ['Type', 'Metric'] + [col for col in combined.columns if col not in ['Type', 'Metric']]
    combined = combined[cols]

    # Sorter etter Type og ønsket Metric-rekkefølge
    combined['Metric'] = pd.Categorical(combined['Metric'], categories=[metric_names[m] for m in metrics], ordered=True)
    combined.sort_values(by=['Type', 'Metric'], inplace=True)
    return combined

def add_demand_to_summary(summary_table, demand_data, group_col):
    """
    Legger til demand_avg som en ekstra rad i summary-tabellen med teksten 'load' i 'Type'-kolonnen.
    """
    # Summér demand_avg for hver gruppe
    demand_summary = demand_data.groupby(group_col)['demand_avg'].sum().reset_index()

    # Opprett en ny rad for Demand Average
    demand_row = {col: "N/A" for col in summary_table.columns}
    demand_row['Type'] = 'load'  # Endret til 'load'
    demand_row['Metric'] = 'Demand Average [MW]'

    # Mapper verdiene til de riktige kolonnene
    for col in summary_table.columns:
        if col in demand_summary[group_col].values:
            demand_row[col] = demand_summary.loc[demand_summary[group_col] == col, 'demand_avg'].values[0]

    # Legg til raden nederst
    summary_table = pd.concat([summary_table, pd.DataFrame([demand_row])], ignore_index=True)
    return summary_table

def process_total_data(generator_data, consumer_data):
    """
    Behandler totaldata for hele systemet og legger til total demand.
    """
    metric_names = {
        'pmax': 'Installed capacity [MW]',
        'fuelcost': 'Fuelcost [Eur/MWh]',
        'storage_cap': 'Storage capacity [TWh]',
        'storage_price': 'Storage price [Eur/MWh]',
        'inflow_fac': 'Inflow factor'
    }

    metrics = ['pmax', 'fuelcost', 'storage_cap', 'storage_price', 'inflow_fac']
    total_data = []

    # Beregn totalsummer for generator-data
    for metric in metrics:
        for fuel_type in generator_data['type'].unique():
            if metric in ['fuelcost', 'storage_price', 'inflow_fac']:
                value = generator_data.loc[generator_data['type'] == fuel_type, metric].mean()
            else:
                value = generator_data.loc[generator_data['type'] == fuel_type, metric].sum()
            total_data.append({
                'Type': fuel_type,
                'Metric': metric_names[metric],
                'System Total': value if pd.notna(value) else 'N/A'
            })

    # Beregn total demand
    total_demand = consumer_data['demand_avg'].sum()
    total_data.append({
        'Type': 'load',
        'Metric': 'Demand Average [MW]',
        'System Total': total_demand
    })

    # Opprett DataFrame og legg til en sorteringskolonne
    gen_total = pd.DataFrame(total_data)
    gen_total['SortOrder'] = gen_total['Type'].apply(lambda x: 1 if x == 'load' else 0)  # Sørg for at 'load' er sist

    gen_total['Metric'] = pd.Categorical(
        gen_total['Metric'],
        categories=[metric_names[m] for m in metrics] + ['Demand Average [MW]'],
        ordered=True
    )
    gen_total.sort_values(by=['SortOrder', 'Type', 'Metric'], inplace=True)
    gen_total.drop(columns=['SortOrder'], inplace=True)  # Fjern SortOrder før lagring
    return gen_total

def save_to_excel(gen_node, gen_zone, gen_country, gen_total, case):
    """
    Lagrer data til Excel med riktig formatering.
    """
    output_path = pathlib.Path(f"../case_{case}/data/case_doc_{case}.xlsx")
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Skriv hvert nivå til sitt eget ark
        sheets = {
            "Node Summary": gen_node,
            "Zone Summary": gen_zone,
            "Country Summary": gen_country,
            "System Total": gen_total
        }
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            format_worksheet(writer.sheets[sheet_name])

    print(f"Data lagret til {output_path}")

def format_worksheet(worksheet):
    """
    Formaterer regnearket: justerer bredder, fet skrift og sammenslåing av celler.
    """
    # Gjør Type-feltet fet og høyrejuster "N/A"
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.col_idx == 1:
                cell.font = Font(bold=True)  # Gjør kolonne A (Type) fet
            if cell.value == "N/A":
                cell.alignment = Alignment(horizontal="right")  # Høyrejuster "N/A"

    # Sammenslå Type-celler for ryddig visning
    current_type, start_row = None, None
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        cell = row[0]
        if cell.value != current_type:
            if current_type is not None and start_row is not None:
                worksheet.merge_cells(start_row=start_row, start_column=1, end_row=row_idx - 1, end_column=1)
            current_type, start_row = cell.value, row_idx
    if current_type is not None and start_row is not None:
        worksheet.merge_cells(start_row=start_row, start_column=1, end_row=row_idx, end_column=1)

    # Juster kolonnebredder for lesbarhet
    for col_idx, col_cells in enumerate(worksheet.iter_cols(), start=1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

def create_case_doc(case):
    """
    Hovedfunksjon for å lese, prosessere og lagre dokumentasjon.
    """
    generator_df, consumer_df = read_data(case)

    # Prosesser generator-data
    gen_node = process_grouped_data(generator_df, 'node')
    generator_df['zone'] = generator_df['node'].str[:3]
    gen_zone = process_grouped_data(generator_df, 'zone')
    generator_df['country'] = generator_df['node'].str[:2]
    gen_country = process_grouped_data(generator_df, 'country')
    gen_total = process_total_data(generator_df, consumer_df)

    # Sørg for at zone og country eksisterer i consumer_df
    consumer_df['zone'] = consumer_df['node'].str[:3]
    consumer_df['country'] = consumer_df['node'].str[:2]

    # Legg demand til tabellene
    gen_node = add_demand_to_summary(gen_node, consumer_df, 'node')
    gen_zone = add_demand_to_summary(gen_zone, consumer_df, 'zone')
    gen_country = add_demand_to_summary(gen_country, consumer_df, 'country')

    # Lagre til Excel
    save_to_excel(gen_node, gen_zone, gen_country, gen_total, case)


# create_case_doc('BM')

