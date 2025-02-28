from fontTools.cffLib import topDictOperators

from functions.global_functions import *
from functions.database_functions import *
from datetime import datetime, timedelta    # For time calculations
from openpyxl import Workbook
from datetime import datetime, timedelta
from pathlib import Path


def calcSystemCostAndMeanPriceFromDB(data: GridData, database: Database, time_SC, time_MP):
    print(f"System cost {sum(getSystemCostFromDB(data=data, db=database, timeMaxMin=time_SC).values()):.2f} EUR, or {sum(getSystemCostFromDB(data=data, db=database, timeMaxMin=time_SC).values())/1e9:.2f} Billion EUR")
    print(f"Mean area price {sum(getAreaPricesAverageFromDB(data=data, db=database, areas=None, timeMaxMin=time_MP).values()) / len(getAreaPricesAverageFromDB(data=data, db=database, areas=None, timeMaxMin=time_MP)):.2f} EUR/MWh")

    
def plot_Map(data: GridData, database: Database, time_Map, DATE_START, OUTPUT_PATH, version):
    correct_date_start = DATE_START + pd.Timedelta(hours=time_Map[0])
    correct_date_end = DATE_START + pd.Timedelta(hours=time_Map[-1])
    output_path = OUTPUT_PATH / f'prices_and_branch_utilization_map_{version}_{correct_date_start.year}_{correct_date_end.year}.html'

    create_price_and_utilization_map_FromDB(data, database, time_max_min=time_Map, output_path=output_path)


def plot_SF_Areas_FromDB(data: GridData, database: Database, time_SF, OUTPUT_PATH_PLOTS, DATE_START, plot_config):
    storfilling = pd.DataFrame()

    for area in plot_config['areas']:
        storfilling[area] = getStorageFillingInAreasFromDB(data=data,
                                                           db=database,
                                                           areas=[area],
                                                           generator_type="hydro",
                                                           relative_storage=plot_config['relative'],
                                                           timeMaxMin=time_SF)
        if plot_config['relative']:
            storfilling[area] = storfilling[area] * 100

    # Compute the correct DATE_START for this year
    correct_date_start_SF = DATE_START + pd.Timedelta(hours=time_SF[0])
    correct_date_end_SF = DATE_START + pd.Timedelta(hours=time_SF[-1])

    storfilling.index = pd.date_range(correct_date_start_SF, periods=time_SF[-1] - time_SF[0], freq='h')
    storfilling['year'] = storfilling.index.year  # Add year column to DataFrame
    title_storage_filling = f"Reservoir Filling in {plot_config['areas']} for period {correct_date_start_SF.year}-{correct_date_end_SF.year}"
    plot_storage_filling_area(storfilling=storfilling,
                              DATE_START=correct_date_start_SF,
                              DATE_END=correct_date_end_SF,
                              areas=plot_config['areas'],
                              interval=plot_config['interval'],
                              title=title_storage_filling,
                              OUTPUT_PATH_PLOTS=OUTPUT_PATH_PLOTS,
                              relative=plot_config['relative'],
                              plot_by_year=plot_config['plot_by_year'],
                              save_plot=plot_config['save_fig'],
                              duration_curve=plot_config['duration_curve'],
                              tex_font=False)



def plot_SF_Zones_FromDB(data: GridData, database: Database, time_SF ,OUTPUT_PATH_PLOTS, DATE_START, plot_config):
    storfilling = pd.DataFrame()
    # TODO: Fix plot by year til å faktisk plotte flere plots for hvert år.

    for zone in plot_config['zones']:
        storfilling[zone] = getStorageFillingInZonesFromDB(data=data,
                                                           db=database,
                                                           zones=[zone],
                                                           generator_type="hydro",
                                                           relative_storage=plot_config['relative'],
                                                           timeMaxMin=time_SF)
        if plot_config['relative']:
            storfilling[zone] = storfilling[zone] * 100

    # Compute the correct DATE_START for this year
    correct_date_start_SF = DATE_START + pd.Timedelta(hours=time_SF[0])
    correct_date_end_SF = DATE_START + pd.Timedelta(hours=time_SF[-1])

    storfilling.index = pd.date_range(correct_date_start_SF, periods=time_SF[-1] - time_SF[0], freq='h')
    storfilling['year'] = storfilling.index.year    # Add year column to DataFrame

    if plot_config['plot_by_year']:
        for year in storfilling['year'].unique():
            title_storage_filling = f"Reservoir Filling in {'Zones: ' + ', '.join(plot_config['zones'])} for year {year}"
            storfilling_year = storfilling[storfilling['year'] == year]
            storfilling_year.index = pd.date_range(correct_date_start_SF, periods=storfilling_year.shape[0], freq='h')
            plot_storage_filling_area(storfilling=storfilling_year,
                                      DATE_START=correct_date_start_SF,
                                      DATE_END=correct_date_end_SF,
                                      areas=plot_config['zones'],
                                      interval=plot_config['interval'],
                                      title=title_storage_filling,
                                      OUTPUT_PATH_PLOTS=OUTPUT_PATH_PLOTS,
                                      relative=plot_config['relative'],
                                      plot_by_year=plot_config['plot_by_year'],
                                      save_plot=plot_config['save_fig'],
                                      duration_curve=plot_config['duration_curve'],
                                      tex_font=False)

    else:
        title_storage_filling = f"Reservoir Filling in {plot_config['zones']} for period {correct_date_start_SF.year}-{correct_date_end_SF.year}"
        plot_storage_filling_area(storfilling=storfilling,
                                  DATE_START=correct_date_start_SF,
                                  DATE_END=correct_date_end_SF,
                                  areas=plot_config['zones'],
                                  interval=plot_config['interval'],
                                  title=title_storage_filling,
                                  OUTPUT_PATH_PLOTS=OUTPUT_PATH_PLOTS,
                                  relative=plot_config['relative'],
                                  plot_by_year=plot_config['plot_by_year'],
                                  save_plot=plot_config['save_fig'],
                                  duration_curve=plot_config['duration_curve'],
                                  tex_font=False)



def calcPlot_NP_FromDB(data: GridData, database: Database, time_NP, OUTPUT_PATH_PLOTS, DATE_START, plot_config):

    nodes_in_zone = data.node[data.node['zone'] == plot_config['zone']].index.tolist() # Get all nodes in the zone
    # Get nodal prices for all nodes in the zone in one step node_prices
    node_prices = pd.DataFrame({node: getNodalPricesFromDB(database, node, time_NP) for node in nodes_in_zone})#  * EUR_MWH_TO_ORE_KWH

    correct_date_start_NP = DATE_START + pd.Timedelta(hours=time_NP[0])
    correct_date_end_NP = DATE_START + pd.Timedelta(hours=time_NP[-1])
    node_prices.index = pd.date_range(correct_date_start_NP, periods=time_NP[-1] - time_NP[0], freq='h')
    title_nodal = f"Avg. Prices in {plot_config['zone']} for period {correct_date_start_NP.year}-{correct_date_end_NP.year}"
    # title_nodal = f"Nodal Prices in {zone} for period {YEAR_START}-{YEAR_END}"
    plot_nodal_prices_FromDB(data=data,
                             node_prices=node_prices,
                             nodes_in_zone=nodes_in_zone,
                             zone=plot_config['zone'],
                             DATE_START=correct_date_start_NP,
                             DATE_END=correct_date_end_NP,
                             interval=plot_config['interval'],
                             TITLE=title_nodal,
                             plot_all_nodes=plot_config['plot_all_nodes'],
                             save_plot_nodal=plot_config['save_fig'],
                             OUTPUT_PATH_PLOTS=OUTPUT_PATH_PLOTS,
                             plot_by_year_nodal=plot_config['plot_by_year'],
                             duration_curve_nodal=plot_config['duration_curve'],
                             tex_font=plot_config['tex_font'])



def calcPlot_ZonalPrices_FromDB(data: GridData, database: Database, time_NP, OUTPUT_PATH_PLOTS, DATE_START, plot_config):
    correct_date_start_NP = DATE_START + pd.Timedelta(hours=time_NP[0])
    correct_date_end_NP = DATE_START + pd.Timedelta(hours=time_NP[-1])

    zonal_prices = pd.DataFrame()

    for zone in plot_config['zones']:
        nodes_in_zone = data.node[data.node['zone'] == zone].index.tolist() # Get all nodes in the zone
        # Get nodal prices for all nodes in the zone in one step node_prices
        node_prices = pd.DataFrame({node: getNodalPricesFromDB(database, node, time_NP) for node in nodes_in_zone})
        node_prices.index = pd.date_range(correct_date_start_NP, periods=time_NP[-1] - time_NP[0], freq='h')
        avg_node_prices = pd.DataFrame((node_prices.sum(axis=1) / len(nodes_in_zone)), columns=[f'avg_price_{zone}'])
        zonal_prices[f'avg_price_{zone}'] = avg_node_prices[f'avg_price_{zone}']



    zonal_prices.index = pd.date_range(correct_date_start_NP, periods=time_NP[-1] - time_NP[0], freq='h')
    zonal_prices[f'year'] = zonal_prices.index.year  # Add year column to DataFrame
    title_zonal = f"Avg. Prices in {'Zones: ' + ', '.join(plot_config['zones'])} for period {correct_date_start_NP.year}-{correct_date_end_NP.year}"
    plot_zonal_prices_FromDB(data=data,
                             zone_prices=zonal_prices,
                             zones=plot_config['zones'],
                             DATE_START=correct_date_start_NP,
                             DATE_END=correct_date_end_NP,
                             interval=plot_config['interval'],
                             TITLE=title_zonal,
                             save_plot_nodal=plot_config['save_fig'],
                             OUTPUT_PATH_PLOTS=OUTPUT_PATH_PLOTS,
                             plot_by_year=plot_config['plot_by_year'],
                             duration_curve_nodal=plot_config['duration_curve'],
                             tex_font=plot_config['tex_font'])


def calcPlot_HRI_FromDB(data: GridData, database: Database, time_HRI, OUTPUT_PATH_PLOTS, DATE_START, plot_config):


    correct_date_start_HRI = DATE_START + pd.Timedelta(hours=time_HRI[0])
    correct_date_end_HRI = DATE_START + pd.Timedelta(hours=time_HRI[-1])

    df_resampled = calculate_Hydro_Res_Inflow_FromDB(data,
                                                     database,
                                                     correct_date_start_HRI,
                                                     plot_config['area'],
                                                     plot_config['genType'],
                                                     time_HRI,
                                                     plot_config['relative_storage'],
                                                     include_pump=False)
    df_resampled['year'] = df_resampled.index.year
    title_HRI = 'Hydro Production, Reservoir Filling and Inflow'

    plot_hydro_prod_res_inflow(df=df_resampled,
                               DATE_START=correct_date_start_HRI,
                               DATE_END=correct_date_end_HRI,
                               interval=plot_config['interval'],
                               TITLE=title_HRI,
                               OUTPUT_PATH_PLOTS=OUTPUT_PATH_PLOTS,
                               save_plot=plot_config['save_fig'],
                               box_in_frame=plot_config['box_in_frame'],
                               plot_full_timeline=plot_config['plot_full_timeline'],
                               tex_font=False)



def calcPlot_PLP_FromDB(data: GridData, database: Database, time_PLP, OUTPUT_PATH_PLOTS, DATE_START, plot_config):

    correct_date_start_PLP = DATE_START + pd.Timedelta(hours=time_PLP[0])
    correct_date_end_PLP = DATE_START + pd.Timedelta(hours=time_PLP[-1])


    df_plp, df_plp_resampled = calc_PLP_FromDB(data, database, plot_config['area'], correct_date_start_PLP, time_PLP)
    plot_hydro_prod_demand_price(df_plp=df_plp,
                                 df_plp_resampled=df_plp_resampled,
                                 resample=plot_config['resample'],
                                 DATE_START=correct_date_start_PLP,
                                 DATE_END=correct_date_end_PLP,
                                 interval=plot_config['interval'],
                                 TITLE=plot_config['title'],
                                 save_fig=plot_config['save_fig'],
                                 plot_full_timeline=plot_config['plot_full_timeline'],
                                 box_in_frame=plot_config['box_in_frame'],
                                 OUTPUT_PATH_PLOTS=OUTPUT_PATH_PLOTS,
                                 tex_font=False)


def calcPlot_LG_FromDB(data: GridData, database: Database, time_LGT, OUTPUT_PATH_PLOTS, DATE_START, plot_config):
    correct_date_start_LGT = DATE_START + pd.Timedelta(hours=time_LGT[0])
    correct_date_end_LGT = DATE_START + pd.Timedelta(hours=time_LGT[-1])

    df_gen_resampled, df_prices_resampled, total_production = get_production_by_type_FromDB(data,
                                                                                            database,
                                                                                            plot_config['area'],
                                                                                            time_LGT,
                                                                                            correct_date_start_LGT)

    plot_production(df_gen_resampled=df_gen_resampled,
                    df_prices_resampled=df_prices_resampled,
                    DATE_START=correct_date_start_LGT,
                    DATE_END=correct_date_end_LGT,
                    interval=plot_config['interval'],
                    fig_size=plot_config['fig_size'],
                    TITLE=plot_config['title'],
                    OUTPUT_PATH_PLOTS=OUTPUT_PATH_PLOTS,
                    plot_full_timeline=plot_config['plot_full_timeline'],
                    plot_duration_curve=plot_config['duration_curve'],
                    save_fig=plot_config['save_fig'],
                    box_in_frame=plot_config['box_in_frame'],
                    tex_font=False)

    return df_gen_resampled, df_prices_resampled, total_production



def getProductionZonesInArea(data: GridData, db: Database, area=None, time_max_min=None, DATE_START=None, week=None):
    """
    Fetches the production by type for all zones within a specified area and exports the result to a CSV file.
    """
    zones_in_area_prod = get_production_by_type_FromDB_ZoneLevel(data, db, area=area, time_max_min=time_max_min,
                                                                 DATE_START=DATE_START, week=week)
    zones_in_area_prod.to_csv(f'production_zone_level_{area}_{DATE_START.year}.csv')
    return zones_in_area_prod


def getProductionNodesInZone(data: GridData, db: Database, zone=None, time_max_min=None, DATE_START=None, week=None):
    """
    Retrieves the production by types for all nodes in a specified zone and exports the result to a CSV file.
    """
    nodes_in_zone_prod = get_production_by_type_FromDB_NodesInZone(data, db, zone=zone, time_max_min=time_max_min,
                                                                   DATE_START=DATE_START, week=week)
    nodes_in_zone_prod.to_csv(f'production_nodes_in_zone_{zone}_{DATE_START.year}.csv')
    return nodes_in_zone_prod

#######################################################################################################################
# EINAR ###############################################################################################################
#######################################################################################################################

# Time handling function using Python's built-in datetime objects.
def get_hour_range(YEAR_START, YEAR_END, TIMEZONE, start, end):
    """
    Beregner timeindeks for et gitt tidsintervall basert på simuleringens starttidspunkt.
    Simuleringen starter ved YEAR_START (Gitt i "General Configurations" ).
    Tar inn to tidspunkt (start og slutt) i dictionary-format og returnerer
    indeksene for disse tidspunktene i forhold til simuleringens start.

    Eksempel:
    YEAR_START: Startåret for simuleringsperioden (SQL)
    YEAR_END: Sluttåret for simuleringsperioden (SQL)
    TIMEZONE = ZoneInfo("UTC")
    START = {"year": 2005, "month": 5, "day": 5, "hour": 14}
    END = {"year": 2005, "month": 5, "day": 8, "hour": 14}
    get_hour_range(START, END)

    Funksjonen håndterer skuddår.
    Gir feilmelding dersom start og slutt ikke er innenfor perioden som SQL-filen har simulert over.
    """

    if not (YEAR_START <= start["year"] <= YEAR_END) or not (YEAR_START <= end["year"] <= YEAR_END):
        raise ValueError(f"Input years must be within {YEAR_START}-{YEAR_END}")

    start_time = datetime(YEAR_START, 1, 1, 0, 0, tzinfo=TIMEZONE)

    # Definer start- og sluttidspunktet basert på input
    start_datetime = datetime(start["year"], start["month"], start["day"], start["hour"], 0, tzinfo=TIMEZONE)
    end_datetime = datetime(end["year"], end["month"], end["day"], end["hour"], 0, tzinfo=TIMEZONE)

    # Beregn timeindeks
    start_hour_index = int((start_datetime - start_time).total_seconds() / 3600)
    end_hour_index = int((end_datetime - start_time).total_seconds() / 3600)

    print(f"Start hour index: {start_hour_index}")
    print(f"End hour index: {end_hour_index}")
    print(f"Time steps: {end_hour_index - start_hour_index}")

    return start_hour_index, end_hour_index


#### Get production in specific node

def GetProductionAtSpecificNodes(Nodes, data, database, start_hour, end_hour):
    """
    Henter produksjonsdata for spesifikke noder i et gitt tidsintervall.

    Args:
        Nodes (list): Liste over nodenavn (f.eks. ['NO1_1', 'NO1_2']).
        data (object): Datastruktur med informasjon om noder, generatorer osv.
        database (object): Databaseforbindelse for å hente produksjonsdata.
        start_hour (int): Startindeks for tidsserien.
        end_hour (int): Sluttindeks for tidsserien.

    Returns:
        tuple:
            - production_per_node (dict): Produksjonsdata per node og type.
            - gen_idx (list): Liste over generator-IDer per node.
            - gen_type (list): Liste over generatortyper per node.
    """

    # === FINN INDEKSENE FOR NODENE ===
    node_idx = [int(data.node[data.node['id'] == node].index[0]) for node in Nodes]

    # === HENT GENERATORER OG DERES TYPER ===
    gen_idx = [[gen for gen in data.getGeneratorsAtNode(idx)] for idx in node_idx]
    gen_type = [[data.generator.loc[gen, "type"] for gen in gens] for gens in gen_idx]
    flat_gen_idx = [gen for sublist in gen_idx for gen in sublist]  # Flater ut listen

    # === HENT PRODUKSJONSDATA FRA DATABASE ===
    power_output = {gen: database.getResultGeneratorPower([gen], (start_hour, end_hour)) for gen in flat_gen_idx}

    # === ORGANISERE PRODUKSJON PER NODE OG TYPE ===
    production_per_node = {node: {} for node in Nodes}
    for node, gen_list, type_list in zip(Nodes, gen_idx, gen_type):
        for gen, typ in zip(gen_list, type_list):
            production_per_node[node].setdefault(typ, []).append(power_output.get(gen, [0]))  # Setter 0 hvis data mangler

    return production_per_node, gen_idx, gen_type



def GetConsumptionAtSpecificNodes(Nodes, data, database, start_hour, end_hour):
    """
    Henter forbruksdata for spesifikke noder i et gitt tidsintervall.

    Args:
        Nodes (list): Liste over nodenavn (f.eks. ['NO1_1', 'NO1_2']).
        data (object): Datastruktur med informasjon om noder.
        database (object): Databaseforbindelse for å hente forbruksdata.
        start_hour (int): Startindeks for tidsserien.
        end_hour (int): Sluttindeks for tidsserien.

    Returns:
        dict: Forbruksdata per node med kategoriene "fixed", "flex" og "sum".
    """

    # === FINN INDEKSENE FOR NODENE ===
    node_idx = [int(data.node[data.node['id'] == node].index[0]) for node in Nodes]

    # === HENT FORBRUKSDATA FOR HVER NODE ===
    consumption_per_node = {node: {} for node in Nodes}

    for node, idx in zip(Nodes, node_idx):
        area = data.node.loc[idx, "area"]  # Finn område for noden
        demand_data = getDemandPerNodeFromDB(data, database, area, node, (start_hour, end_hour))

        consumption_per_node[node]["fixed"] = demand_data["fixed"]
        consumption_per_node[node]["flex"] = demand_data["flex"]
        consumption_per_node[node]["sum"] = demand_data["sum"]

    return consumption_per_node


def GetPriceAtSpecificNodes(Nodes, data, database, start_hour, end_hour):
    """
    Henter nodalpris for spesifikke noder i et gitt tidsintervall.

    Args:
        Nodes (list): Liste over nodenavn.
        data (object): Datastruktur med informasjon om noder.
        database (object): Databaseforbindelse for å hente priser.
        start_hour (int): Startindeks for tidsserien.
        end_hour (int): Sluttindeks for tidsserien.

    Returns:
        dict: Nodalpris per node.
    """

    # === FINN INDEKSENE FOR NODENE ===
    node_idx = [int(data.node[data.node['id'] == node].index[0]) for node in Nodes]

    # === HENT NODALPRIS FOR HVER NODE ===
    nodal_prices = {node: database.getResultNodalPrice(idx, (start_hour, end_hour)) for node, idx in zip(Nodes, node_idx)}

    return nodal_prices




def ExportToExcel(Nodes, production_per_node, consumption_per_node, nodal_prices_per_node, reservoir_filling_per_node, storage_cap, START, END, case, version, OUTPUT_PATH):
    """
    Eksporterer produksjons-, forbruks-, fyllingsgrads- og nodalprisdata til en Excel-fil.

    Args:
        Nodes (list): Liste over nodenavn.
        production_per_node (dict): Produksjonsdata per node og type.
        consumption_per_node (dict): Forbruksdata per node.
        nodal_prices_per_node (dict): Nodalpriser per node.
        reservoir_filling_per_node (dict): Reservoarfylling per node.
        START (dict): Starttidspunkt som dictionary (f.eks. {"year": 2019, "month": 5, "day": 1, "hour": 12}).
        END (dict): Sluttidspunkt som dictionary (f.eks. {"year": 2019, "month": 6, "day": 1, "hour": 12}).
        case (str): Navn på caset.
        version (str): Versjonsnummer.
        OUTPUT_PATH: Filsti for lagring.

    Returns:
        str: Filnavn på den lagrede Excel-filen.
    """

    # Konverter START og END til datetime-objekter
    start_datetime = datetime(START["year"], START["month"], START["day"], START["hour"])
    end_datetime = datetime(END["year"], END["month"], END["day"], END["hour"])

    all_types = ["nuclear", "hydro", "biomass", "ror", "wind_on", "wind_off", "solar", "fossile_other", "fossile_gas"]

    # === GENERER TIDSSTEG BASERT PÅ get_hour_range() ===
    time_stamps = [start_datetime + timedelta(hours=i) for i in range(int((end_datetime - start_datetime).total_seconds() // 3600) + 1)]

    # === GENERER FILNAVN ===
    timestamp = datetime.now().strftime("%Y-%m-%d")
    start_str = start_datetime.strftime("%Y-%m-%d-%H")
    end_str = end_datetime.strftime("%Y-%m-%d-%H")
    filename = f"Prod_demand_nodes_{case}_{version}_{start_str}_to_{end_str}.xlsx"

    # === OPPRETT NY WORKBOOK ===
    wb = Workbook()

    for node in Nodes:
        # === OPPRETT ARKFANER FOR HVER NODE ===
        ws_production = wb.create_sheet(f"Production {node}")
        ws_consumption = wb.create_sheet(f"Consumption {node}")
        ws_price = wb.create_sheet(f"Price {node}")
        ws_reservoir = wb.create_sheet(f"Reservoir {node}")

        # === LEGG TIL OVERSKRIFTER ===
        ws_production.append(["Timestamp"] + all_types)
        ws_consumption.append(["Timestamp", "Fixed", "Flexible", "Consumption"])
        ws_price.append(["Timestamp", "Nodal Price"])
        ws_reservoir.append(["Timestamp", "Reservoir Filling", "Max storage capacity", "Reservoir Filling [%]"])


        # === FYLL PRODUKSJONSARKET ===
        for t, timestamp in enumerate(time_stamps):
            row = [timestamp.strftime("%Y-%m-%d %H:%M")]  # Formater tid riktig
            for typ in all_types:
                values = production_per_node[node].get(typ, [[0]])[0]  # Fjern dobbel liste-nesting
                value = values[t] if t < len(values) else 0  # Hent riktig indeks eller sett 0
                row.append(value)
            ws_production.append(row)

        # === FYLL FORBRUKSARKET ===
        for t, timestamp in enumerate(time_stamps):
            fixed = consumption_per_node[node]["fixed"]
            flex = consumption_per_node[node]["flex"]
            total = consumption_per_node[node]["sum"]

            ws_consumption.append([
                timestamp.strftime("%Y-%m-%d %H:%M"),
                fixed[t] if t < len(fixed) else 0,
                flex[t] if t < len(flex) else 0,
                total[t] if t < len(total) else 0,
            ])

        # === FYLL NODALPRISARKET ===
        for t, timestamp in enumerate(time_stamps):
            nodal_price = nodal_prices_per_node[node]  # Hent liste over nodalpriser for noden
            price_value = nodal_price[t] if t < len(nodal_price) else 0  # Hent pris eller sett 0 hvis ikke nok data

            ws_price.append([
                timestamp.strftime("%Y-%m-%d %H:%M"),
                price_value,
            ])

        # === FYLL RESERVOARFYLLINGSARKET ===

        for t, timestamp in enumerate(time_stamps):
            # Henter og pakker ut reservoardata
            if node in reservoir_filling_per_node:
                reservoir_data = reservoir_filling_per_node.get(node, [])
                if isinstance(reservoir_data, list):  # Sikrer riktig format
                    reservoir_values = [sum(x) for x in zip(*reservoir_data)]  # Slår sammen flere generatorer
                else:
                    reservoir_values = [0] * len(time_stamps)  # Hvis data er feil format
            else:
                reservoir_values = [0] * len(time_stamps)  # Hvis noden ikke finnes

            # Hent riktig verdi fra tidsserien eller sett 0
            reservoir_value = reservoir_values[t] if t < len(reservoir_values) else 0

            # Hent maks kapasitet for noden
            max_capacity = storage_cap.get(node, 0)  # Standard 0 hvis ikke funnet

            # Beregn fyllingsgrad (%)
            filling_percentage = (reservoir_value / max_capacity) * 100 if max_capacity > 0 else 0

            ws_reservoir.append([
                timestamp.strftime("%Y-%m-%d %H:%M"),
                round(reservoir_value, 4),  # Rund av til 4 desimaler
                round(max_capacity, 4) if max_capacity > 0 else "",  # Tom celle hvis ikke kapasitet
                round(filling_percentage, 8)  # Rund av til 8 desimaler
            ])


    # Fjern default ark
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # === LAGRE FIL ===
    filepath = Path(OUTPUT_PATH) / filename
    wb.save(filepath)

    print(f"\nExcel-fil '{filename}' er lagret i {OUTPUT_PATH}!")

    return filename


def GetReservoirFillingAtSpecificNodes(Nodes, data, database, start_hour, end_hour):
    """
    Henter reservoarfylling og maksimal kapasitet for spesifikke noder.
    """

    # === FINN INDEKSENE FOR NODENE ===
    node_idx = [int(data.node[data.node['id'] == node].index[0]) for node in Nodes]

    # === HENT LAGRINGSENHETER OG DERES KAPASITET ===
    storage_data = data.generator[
        (data.generator["node"].isin(Nodes)) &
        (data.generator["storage_cap"] > 0) &
        (data.generator["type"] == "hydro")
    ][["node", "storage_cap"]]

    storage_idx = storage_data.groupby("node").apply(lambda x: list(x.index)).to_dict()
    storage_cap = storage_data.groupby("node")["storage_cap"].sum().to_dict()  # Summerer kapasitet for hver node

    flat_storage_idx = [gen for sublist in storage_idx.values() for gen in sublist]

    # === HENT RESERVOARFYLLINGSNIVÅ FRA DATABASE ===
    storage_filling = {gen: database.getResultStorageFilling(gen, (start_hour, end_hour)) for gen in flat_storage_idx}

    # === ORGANISERE DATA ===
    reservoir_filling_per_node = {node: [] for node in Nodes}

    for node, gen_list in storage_idx.items():
        node_values = []
        for gen in gen_list:
            node_values.append(storage_filling.get(gen, [0]))  # Hent fyllingsdata eller 0
        reservoir_filling_per_node[node] = node_values

    return reservoir_filling_per_node, storage_cap





